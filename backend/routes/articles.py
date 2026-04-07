"""Article endpoints (public + admin), clusters, categories, OG tags, sitemap, bulk import."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from typing import Optional, List
from datetime import datetime, timezone
import io
import uuid

from core import db, verify_admin, SITE_URL
from models import Article, ArticleCreate, ArticleUpdate

router = APIRouter()

# Lightweight fields for blog list
BLOG_LIST_FIELDS = {
    "_id": 0, "id": 1, "slug": 1, "title": 1, "excerpt": 1,
    "image": 1, "category": 1, "cluster": 1, "date": 1,
    "readTime": 1, "featured": 1, "author": 1
}


@router.get("/articles/list")
async def get_articles_list(
    cluster: Optional[str] = None,
    category: Optional[str] = None,
    page: int = 1,
    limit: int = 12,
    search: Optional[str] = None,
    lang: Optional[str] = None
):
    """Get paginated article list for blog page (lightweight - no content)"""
    query = {}
    if cluster and cluster != "All":
        query["category"] = cluster
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"excerpt": {"$regex": search, "$options": "i"}}
        ]

    total = await db.articles.count_documents(query)

    skip = (page - 1) * limit
    articles = await db.articles.find(query, BLOG_LIST_FIELDS).sort([("sortOrder", 1), ("date", -1)]).skip(skip).limit(limit).to_list(limit)

    # Translate titles and excerpts if lang=en
    if lang == "en":
        from routes.translate import _get_or_translate
        for art in articles:
            if art.get("title"):
                art["title"] = await _get_or_translate(art["title"], "de", "en")
            if art.get("excerpt"):
                art["excerpt"] = await _get_or_translate(art["excerpt"], "de", "en")
            if art.get("category"):
                art["category"] = await _get_or_translate(art["category"], "de", "en")

    return {
        "articles": articles,
        "total": total,
        "page": page,
        "limit": limit,
        "totalPages": (total + limit - 1) // limit,
        "hasMore": skip + len(articles) < total
    }


@router.get("/articles", response_model=List[Article])
async def get_articles(
    cluster: Optional[str] = None,
    featured: Optional[bool] = None,
    category: Optional[str] = None,
    limit: Optional[int] = None
):
    """Get all articles with optional filters (full data)"""
    query = {}
    if cluster:
        query["cluster"] = cluster
    if featured is not None:
        query["featured"] = featured
    if category:
        query["category"] = category

    cursor = db.articles.find(query, {"_id": 0}).sort([("sortOrder", 1), ("date", -1)])
    if limit:
        cursor = cursor.limit(limit)

    articles = await cursor.to_list(1000)
    return articles


@router.get("/articles/{slug}", response_model=Article)
async def get_article_by_slug(slug: str):
    """Get a single article by slug"""
    article = await db.articles.find_one({"slug": slug}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    return article


@router.get("/og/blog/{slug}")
async def get_article_og_html(slug: str):
    """Serve pre-rendered HTML with OG tags for social media crawlers"""
    article = await db.articles.find_one({"slug": slug}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")

    title = f"{article.get('title', '')} | EuroAdria Corporate Solutions"
    description = article.get('excerpt', '')[:200]
    image = article.get('image', f"{SITE_URL}/euroadria-logo.png")
    url = f"{SITE_URL}/blog/{slug}"

    html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<meta name="description" content="{description}"/>
<meta property="og:type" content="article"/>
<meta property="og:url" content="{url}"/>
<meta property="og:title" content="{title}"/>
<meta property="og:description" content="{description}"/>
<meta property="og:image" content="{image}"/>
<meta property="og:image:width" content="1200"/>
<meta property="og:image:height" content="630"/>
<meta property="og:locale" content="de_DE"/>
<meta property="og:site_name" content="EuroAdria Corporate Solutions"/>
<meta name="twitter:card" content="summary_large_image"/>
<meta name="twitter:title" content="{title}"/>
<meta name="twitter:description" content="{description}"/>
<meta name="twitter:image" content="{image}"/>
<meta http-equiv="refresh" content="0;url={url}"/>
</head>
<body><p>Weiterleitung zu <a href="{url}">{title}</a></p></body>
</html>"""
    return Response(content=html, media_type="text/html")


@router.get("/articles/id/{article_id}", response_model=Article)
async def get_article_by_id(article_id: int):
    """Get a single article by ID"""
    article = await db.articles.find_one({"id": article_id}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    return article


@router.get("/clusters")
async def get_clusters():
    """Get all unique categories with article counts"""
    pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    clusters = await db.articles.aggregate(pipeline).to_list(100)
    return [{"id": c["_id"], "name": c["_id"], "count": c["count"]} for c in clusters if c["_id"]]


@router.get("/categories")
async def get_categories():
    """Get all unique categories with article counts"""
    pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    categories = await db.articles.aggregate(pipeline).to_list(100)
    return [{"category": c["_id"], "count": c["count"]} for c in categories]


# ── Admin Article CRUD ──────────────────────────────────────────────────

@router.post("/admin/articles", response_model=Article)
async def create_article(article: ArticleCreate, admin: str = Depends(verify_admin)):
    """Create a new article (Admin only)"""
    existing = await db.articles.find_one({"slug": article.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Article with this slug already exists")

    max_id_doc = await db.articles.find_one(sort=[("id", -1)])
    new_id = (max_id_doc["id"] + 1) if max_id_doc else 101

    article_dict = article.model_dump()
    article_dict["id"] = new_id

    await db.articles.insert_one(article_dict)
    return article_dict


@router.put("/admin/articles/{article_id}", response_model=Article)
async def update_article(article_id: int, article_update: ArticleUpdate, admin: str = Depends(verify_admin)):
    """Update an article (Admin only)"""
    existing = await db.articles.find_one({"id": article_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Article not found")

    update_data = {k: v for k, v in article_update.model_dump().items() if v is not None}

    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    if "slug" in update_data and update_data["slug"] != existing.get("slug"):
        slug_exists = await db.articles.find_one({"slug": update_data["slug"], "id": {"$ne": article_id}})
        if slug_exists:
            raise HTTPException(status_code=400, detail="Article with this slug already exists")

    await db.articles.update_one({"id": article_id}, {"$set": update_data})

    updated = await db.articles.find_one({"id": article_id}, {"_id": 0})
    return updated


@router.delete("/admin/articles/{article_id}")
async def delete_article(article_id: int, admin: str = Depends(verify_admin)):
    """Delete an article (Admin only)"""
    result = await db.articles.delete_one({"id": article_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    return {"message": "Article deleted successfully", "id": article_id}


# ── Seed & Bulk Import ──────────────────────────────────────────────────

@router.post("/admin/seed-articles")
async def seed_articles(admin: str = Depends(verify_admin)):
    """Seed the database with initial articles (Admin only)"""
    count = await db.articles.count_documents({})
    if count > 0:
        return {"message": f"Database already contains {count} articles. Skipping seed.", "seeded": False}
    return {"message": "Use POST /api/admin/seed-articles-data with articles array to seed", "seeded": False}


@router.post("/admin/seed-articles-data")
async def seed_articles_data(articles: List[ArticleCreate], admin: str = Depends(verify_admin)):
    """Seed the database with provided articles data (Admin only)"""
    await db.articles.delete_many({})

    for i, article in enumerate(articles):
        article_dict = article.model_dump()
        article_dict["id"] = 101 + i
        await db.articles.insert_one(article_dict)

    return {"message": f"Successfully seeded {len(articles)} articles", "seeded": True, "count": len(articles)}


@router.post("/admin/articles/bulk-import")
async def bulk_import_articles(file: UploadFile = File(...), admin: str = Depends(verify_admin)):
    """Bulk import articles from CSV, XLSX, or DOCX files"""
    import csv as csv_module

    filename = file.filename.lower()
    content = await file.read()

    articles_data = []

    try:
        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv_module.DictReader(io.StringIO(text), delimiter=';')
            if not reader.fieldnames or 'Titel' not in [f.strip() for f in reader.fieldnames]:
                reader = csv_module.DictReader(io.StringIO(text), delimiter=',')
            for row in reader:
                row = {k.strip(): v.strip() if v else '' for k, v in row.items() if k}
                if row.get('Titel'):
                    articles_data.append({
                        'title': row.get('Titel', ''),
                        'category': row.get('Kategorie', 'Allgemein'),
                        'excerpt': row.get('Excerpt', row.get('Beschreibung', '')),
                        'content': row.get('Content', row.get('Inhalt', '')),
                        'image': row.get('Bild', row.get('Bild-URL', row.get('Image', ''))),
                        'author': row.get('Autor', 'EuroAdria Corporate Solutions'),
                        'readTime': row.get('Lesezeit', '5 min'),
                        'downloadUrl': row.get('Download-URL', row.get('DownloadUrl', None)) or None,
                    })

        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Excel-Datei ist leer")
            headers = [str(h).strip() if h else '' for h in rows[0]]
            col_map = {}
            for i, h in enumerate(headers):
                hl = h.lower()
                if 'titel' in hl or 'title' in hl: col_map['title'] = i
                elif 'kategorie' in hl or 'category' in hl or 'cluster' in hl: col_map['category'] = i
                elif 'excerpt' in hl or 'beschreibung' in hl: col_map['excerpt'] = i
                elif 'content' in hl or 'inhalt' in hl: col_map['content'] = i
                elif 'bild' in hl or 'image' in hl: col_map['image'] = i
                elif 'autor' in hl or 'author' in hl: col_map['author'] = i
                elif 'lesezeit' in hl or 'readtime' in hl: col_map['readTime'] = i
                elif 'download' in hl: col_map['downloadUrl'] = i

            if 'title' not in col_map:
                raise HTTPException(status_code=400, detail="Spalte 'Titel' nicht gefunden. Erwartete Spalten: Titel, Kategorie, Excerpt, Content, Bild")

            for row in rows[1:]:
                title = str(row[col_map['title']]).strip() if col_map.get('title') is not None and row[col_map['title']] else ''
                if title:
                    articles_data.append({
                        'title': title,
                        'category': str(row[col_map.get('category', 0)] or 'Allgemein').strip() if col_map.get('category') is not None else 'Allgemein',
                        'excerpt': str(row[col_map.get('excerpt', 0)] or '').strip() if col_map.get('excerpt') is not None else '',
                        'content': str(row[col_map.get('content', 0)] or '').strip() if col_map.get('content') is not None else '',
                        'image': str(row[col_map.get('image', 0)] or '').strip() if col_map.get('image') is not None else '',
                        'author': str(row[col_map.get('author', 0)] or 'EuroAdria Corporate Solutions').strip() if col_map.get('author') is not None else 'EuroAdria Corporate Solutions',
                        'readTime': str(row[col_map.get('readTime', 0)] or '5 min').strip() if col_map.get('readTime') is not None else '5 min',
                        'downloadUrl': str(row[col_map.get('downloadUrl', 0)] or '').strip() if col_map.get('downloadUrl') is not None else None,
                    })
            wb.close()

        elif filename.endswith('.docx'):
            from docx import Document
            doc = Document(io.BytesIO(content))
            current_article = None
            current_content = []
            current_category = 'Allgemein'

            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    if current_content:
                        current_content.append('')
                    continue

                if para.style.name.startswith('Heading 1') or (para.runs and para.runs[0].bold and len(text) < 150 and not text.endswith('.')):
                    if current_article:
                        content_html = ''
                        for line in current_content:
                            if line == '':
                                continue
                            elif line.startswith('##'):
                                content_html += f'<h3>{line[2:].strip()}</h3>'
                            else:
                                content_html += f'<p>{line}</p>'
                        current_article['content'] = content_html
                        current_article['excerpt'] = current_content[0][:200] if current_content else ''
                        articles_data.append(current_article)

                    if text.startswith('[') and ']' in text:
                        current_category = text[1:text.index(']')].strip()
                        title_part = text[text.index(']')+1:].strip()
                        if title_part:
                            current_article = {'title': title_part, 'category': current_category, 'content': '', 'excerpt': '', 'image': '', 'author': 'EuroAdria Corporate Solutions', 'readTime': '5 min', 'downloadUrl': None}
                        else:
                            current_article = None
                    else:
                        current_article = {'title': text, 'category': current_category, 'content': '', 'excerpt': '', 'image': '', 'author': 'EuroAdria Corporate Solutions', 'readTime': '5 min', 'downloadUrl': None}
                    current_content = []

                elif para.style.name.startswith('Heading 2'):
                    current_content.append(f'##{text}')
                else:
                    current_content.append(text)

            if current_article:
                content_html = ''
                for line in current_content:
                    if line == '':
                        continue
                    elif line.startswith('##'):
                        content_html += f'<h3>{line[2:].strip()}</h3>'
                    else:
                        content_html += f'<p>{line}</p>'
                current_article['content'] = content_html
                current_article['excerpt'] = current_content[0][:200] if current_content else ''
                articles_data.append(current_article)
        else:
            raise HTTPException(status_code=400, detail="Nicht unterstütztes Format. Bitte CSV, XLSX oder DOCX verwenden.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fehler beim Parsen der Datei: {str(e)}")

    if not articles_data:
        raise HTTPException(status_code=400, detail="Keine Artikel in der Datei gefunden.")

    max_id_doc = await db.articles.find_one(sort=[("id", -1)])
    next_id = (max_id_doc["id"] + 1) if max_id_doc else 101

    import re
    def make_slug(title):
        slug = title.lower().strip()
        slug = slug.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
        slug = re.sub(r'[^a-z0-9]+', '-', slug)
        slug = slug.strip('-')
        return slug[:80]

    existing_articles = await db.articles.find({}, {"slug": 1, "title": 1, "_id": 0}).to_list(10000)
    existing_slugs = {a["slug"] for a in existing_articles}
    existing_titles = {a["title"].strip().lower() for a in existing_articles if a.get("title")}

    imported = []
    skipped = []
    file_seen_slugs = set()

    for art in articles_data:
        title = art['title'].strip()
        if not title:
            skipped.append({"title": "(leer)", "reason": "Kein Titel"})
            continue

        slug = make_slug(title)

        if slug in existing_slugs:
            skipped.append({"title": title, "reason": f"Slug '{slug}' existiert bereits in der Datenbank"})
            continue

        if title.strip().lower() in existing_titles:
            skipped.append({"title": title, "reason": "Titel existiert bereits in der Datenbank"})
            continue

        if slug in file_seen_slugs:
            skipped.append({"title": title, "reason": "Duplikat innerhalb der Import-Datei"})
            continue

        file_seen_slugs.add(slug)

        category = art.get('category', 'Allgemein')

        article_doc = {
            "id": next_id,
            "cluster": category,
            "title": title,
            "slug": slug,
            "excerpt": art.get('excerpt', '')[:300],
            "content": art.get('content', ''),
            "image": art.get('image', ''),
            "category": category,
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "readTime": art.get('readTime', '5 min'),
            "featured": False,
            "author": art.get('author', 'EuroAdria Corporate Solutions'),
            "relatedArticles": [],
            "dueDiligenceBox": None,
            "expertTip": None,
            "downloadUrl": art.get('downloadUrl') or None,
            "sortOrder": 0,
            "imagePosition": 50
        }

        await db.articles.insert_one(article_doc)
        existing_slugs.add(slug)
        existing_titles.add(title.strip().lower())
        imported.append({"id": next_id, "title": title, "category": category, "slug": slug})
        next_id += 1

    result = {
        "message": f"{len(imported)} Artikel importiert, {len(skipped)} übersprungen",
        "count": len(imported),
        "skipped_count": len(skipped),
        "articles": imported,
    }
    if skipped:
        result["skipped"] = skipped

    return result


# ── Dynamic Sitemap ─────────────────────────────────────────────────────

STATIC_PAGES = [
    {"loc": "/", "priority": "1.0", "changefreq": "weekly"},
    {"loc": "/blog", "priority": "0.9", "changefreq": "daily"},
    {"loc": "/serbia-executive", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/serbia-executive/crypto-banking", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/serbia-executive/crypto-compliance", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/infrastruktur-radar", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/investment", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/investment/rechner", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/investment/simulation", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/investment/vergleich", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/contact", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/team", "priority": "0.7", "changefreq": "monthly"},
    {"loc": "/leistungen", "priority": "0.9", "changefreq": "monthly"},
    {"loc": "/events", "priority": "0.7", "changefreq": "weekly"},
    {"loc": "/immobilien/budva", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/niksic", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/podgorica", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/skadar-lake", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/zabljak", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/impressum", "priority": "0.3", "changefreq": "yearly"},
    {"loc": "/datenschutz", "priority": "0.3", "changefreq": "yearly"},
    {"loc": "/agb", "priority": "0.3", "changefreq": "yearly"},
]


@router.get("/sitemap.xml")
async def generate_sitemap():
    """Generate dynamic sitemap from database articles + static pages"""
    today = datetime.now().strftime("%Y-%m-%d")

    xml_parts = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml_parts.append('<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">')

    for page in STATIC_PAGES:
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}{page["loc"]}</loc>')
        xml_parts.append(f'    <lastmod>{today}</lastmod>')
        xml_parts.append(f'    <changefreq>{page["changefreq"]}</changefreq>')
        xml_parts.append(f'    <priority>{page["priority"]}</priority>')
        xml_parts.append('  </url>')

    articles = await db.articles.find({}, {"_id": 0, "slug": 1, "date": 1}).to_list(1000)
    for article in articles:
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}/blog/{article["slug"]}</loc>')
        xml_parts.append(f'    <lastmod>{article.get("date", today)}</lastmod>')
        xml_parts.append('    <changefreq>monthly</changefreq>')
        xml_parts.append('    <priority>0.7</priority>')
        xml_parts.append('  </url>')

    locations = await db.locations.find({}, {"_id": 0, "city": 1}).to_list(100)
    for loc in locations:
        city_slug = loc["city"].replace(" ", "%20")
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}/investment/standort/{city_slug}</loc>')
        xml_parts.append(f'    <lastmod>{today}</lastmod>')
        xml_parts.append('    <changefreq>weekly</changefreq>')
        xml_parts.append('    <priority>0.7</priority>')
        xml_parts.append('  </url>')

    xml_parts.append('</urlset>')

    return Response(
        content="\n".join(xml_parts),
        media_type="application/xml",
        headers={"Content-Type": "application/xml; charset=utf-8"}
    )
