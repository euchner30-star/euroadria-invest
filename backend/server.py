from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, BackgroundTasks, UploadFile, File
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone
import secrets
import io
import resend
import requests as http_requests
import time
import json as json_module
import hashlib


ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR.parent / "frontend" / "public" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

load_dotenv(ROOT_DIR / '.env')

# Object Storage
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "euroadria"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# MongoDB connection (optimized pool for Render Free Tier)
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url, maxPoolSize=10, minPoolSize=1, serverSelectionTimeoutMS=5000)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Simple HTTP Basic Auth for Admin
security = HTTPBasic()

# Admin credentials (in production, use environment variables and proper hashing)
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'euroadria2025')

# Email notification settings
NOTIFICATION_EMAIL = os.environ.get('NOTIFICATION_EMAIL', 'office@euroadria.me')
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '')
BREVO_LIST_ID = 3  # EuroAdria Newsletter list

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# Health check endpoint (keeps Render server awake + verifies DB)
@api_router.get("/health")
async def health_check():
    try:
        await client.admin.command('ping')
        return {"status": "ok", "db": "connected"}
    except Exception:
        return {"status": "ok", "db": "reconnecting"}


def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    """Verify admin credentials"""
    correct_username = secrets.compare_digest(credentials.username, ADMIN_USERNAME)
    correct_password = secrets.compare_digest(credentials.password, ADMIN_PASSWORD)
    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


# Define Models
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str


# Article Models
class DueDiligenceBox(BaseModel):
    title: str
    content: str

class ExpertTip(BaseModel):
    author: str
    title: str
    content: str

class ArticleBase(BaseModel):
    cluster: str
    title: str
    slug: str
    excerpt: str
    content: str
    image: str
    category: str
    date: str
    readTime: str
    featured: bool = False
    author: str
    relatedArticles: List[int] = []
    dueDiligenceBox: Optional[DueDiligenceBox] = None
    expertTip: Optional[ExpertTip] = None
    downloadUrl: Optional[str] = None
    sortOrder: int = 0
    imagePosition: Optional[int] = 50

class ArticleCreate(ArticleBase):
    pass

class ArticleUpdate(BaseModel):
    cluster: Optional[str] = None
    title: Optional[str] = None
    slug: Optional[str] = None
    excerpt: Optional[str] = None
    content: Optional[str] = None
    image: Optional[str] = None
    category: Optional[str] = None
    date: Optional[str] = None
    readTime: Optional[str] = None
    featured: Optional[bool] = None
    author: Optional[str] = None
    relatedArticles: Optional[List[int]] = None
    dueDiligenceBox: Optional[DueDiligenceBox] = None
    expertTip: Optional[ExpertTip] = None
    metaTitle: Optional[str] = None
    metaDescription: Optional[str] = None
    downloadUrl: Optional[str] = None
    sortOrder: Optional[int] = None
    imagePosition: Optional[int] = None

class Article(ArticleBase):
    model_config = ConfigDict(extra="ignore")
    id: int


# Comment Models
class CommentCreate(BaseModel):
    articleId: int
    articleSlug: str
    name: str
    email: str
    content: str

class CommentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    articleId: int
    articleSlug: str
    name: str
    content: str
    status: str
    createdAt: str

class CommentAdminResponse(CommentResponse):
    email: str
    articleTitle: Optional[str] = None


# =============================================
# EVENT MODELS
# =============================================

class EventCreate(BaseModel):
    title: str
    description: str
    date: str  # ISO date string e.g. "2026-05-15"
    time: Optional[str] = None  # e.g. "18:00"
    location: Optional[str] = None
    type: str = "Event"  # Event, Webinar, Workshop
    image: Optional[str] = None
    link: Optional[str] = None  # External registration link
    status: str = "upcoming"  # upcoming, past, cancelled

class EventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    time: Optional[str] = None
    location: Optional[str] = None
    type: Optional[str] = None
    image: Optional[str] = None
    link: Optional[str] = None
    status: Optional[str] = None


# =============================================
# CRM / PIPELINE MODELS
# =============================================

DEFAULT_PIPELINE_STAGES = [
    {"id": "new_lead", "name": "Neuer Lead", "order_index": 1, "probability": 10, "color": "#6B7280"},
    {"id": "qualified", "name": "Qualifiziert", "order_index": 2, "probability": 20, "color": "#3B82F6"},
    {"id": "call_scheduled", "name": "Termin vereinbart", "order_index": 3, "probability": 30, "color": "#8B5CF6"},
    {"id": "consultation_done", "name": "Erstgespräch", "order_index": 4, "probability": 40, "color": "#F59E0B"},
    {"id": "offer_sent", "name": "Angebot gesendet", "order_index": 5, "probability": 60, "color": "#F97316"},
    {"id": "negotiation", "name": "Verhandlung", "order_index": 6, "probability": 80, "color": "#EF4444"},
    {"id": "won", "name": "Gewonnen", "order_index": 7, "probability": 100, "color": "#10B981"},
    {"id": "lost", "name": "Verloren", "order_index": 8, "probability": 0, "color": "#374151"},
]

class CRMLeadCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    lead_source: str = "direct"
    utm_source: Optional[str] = None
    utm_medium: Optional[str] = None
    utm_campaign: Optional[str] = None
    entry_page: Optional[str] = None
    tool_used: Optional[str] = None

class CRMLeadUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    lead_source: Optional[str] = None
    status: Optional[str] = None

class DealCreate(BaseModel):
    lead_id: str
    stage: str = "new_lead"
    deal_value: float = 0
    assigned_to: Optional[str] = None
    notes: Optional[str] = None

class DealUpdate(BaseModel):
    stage: Optional[str] = None
    deal_value: Optional[float] = None
    assigned_to: Optional[str] = None
    notes: Optional[str] = None

class TeamMember(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4())[:8])
    name: str
    title: str
    description: str
    image: str
    email: Optional[str] = None
    phone: Optional[str] = None

class InfoCard(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4())[:8])
    title: str
    description: str
    icon: Optional[str] = None  # Icon name for frontend

class HeroSection(BaseModel):
    title: str
    subtitle: str
    description: Optional[str] = None
    ctaText: Optional[str] = None
    ctaLink: Optional[str] = None
    backgroundImage: Optional[str] = None

class PageSection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4())[:8])
    type: str  # hero, team, cards, text, gallery, cta, testimonials
    title: Optional[str] = None
    content: Optional[str] = None  # WYSIWYG HTML
    data: Optional[dict] = None  # Flexible data for different section types

class PageBase(BaseModel):
    slug: str  # URL path: home, team, contact, etc.
    title: str  # Page title
    metaTitle: Optional[str] = None  # SEO title
    metaDescription: Optional[str] = None  # SEO description
    sections: List[PageSection] = []

class PageCreate(PageBase):
    pass

class PageUpdate(BaseModel):
    title: Optional[str] = None
    metaTitle: Optional[str] = None
    metaDescription: Optional[str] = None
    sections: Optional[List[PageSection]] = None

class Page(PageBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    updatedAt: Optional[str] = None


# Region Models for Admin CMS
class RegionBulletPoint(BaseModel):
    text: str

class RegionApartment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    price: str
    size: str
    imageUrl: str
    features: List[str] = []
    isAvailable: bool = True

class RegionBase(BaseModel):
    slug: str  # e.g., "skadar-lake", "podgorica"
    title: str
    subtitle: str
    investmentScore: int  # 0-100
    priceRange: str  # e.g., "€1.500-3.000/m²"
    potential: str  # e.g., "+35-55%"
    timeHorizon: str  # e.g., "2-5 Jahre"
    content: str  # WYSIWYG HTML content
    bulletPoints: List[RegionBulletPoint] = []
    imageUrls: List[str] = []
    apartments: List[RegionApartment] = []

class RegionCreate(RegionBase):
    pass

class RegionUpdate(BaseModel):
    title: Optional[str] = None
    subtitle: Optional[str] = None
    investmentScore: Optional[int] = None
    priceRange: Optional[str] = None
    potential: Optional[str] = None
    timeHorizon: Optional[str] = None
    content: Optional[str] = None
    bulletPoints: Optional[List[RegionBulletPoint]] = None
    imageUrls: Optional[List[str]] = None
    apartments: Optional[List[RegionApartment]] = None

class Region(RegionBase):
    model_config = ConfigDict(extra="ignore")
    id: str


# Email notification function
async def send_notification_email(comment_data: dict, article_title: str):
    """Send email notification for new comment using Resend"""
    if not RESEND_API_KEY:
        logger.warning("Resend API key not configured - skipping comment email notification")
        return
    
    try:
        resend.api_key = RESEND_API_KEY
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; background-color: #002147; color: #ffffff; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #001233; padding: 30px; border-radius: 10px; border: 1px solid #D4AF37;">
                <h2 style="color: #D4AF37; margin-bottom: 20px;">Neuer Kommentar eingegangen</h2>
                
                <p style="color: #ffffff; margin-bottom: 10px;"><strong>Artikel:</strong> {article_title}</p>
                <p style="color: #ffffff; margin-bottom: 10px;"><strong>Name:</strong> {comment_data['name']}</p>
                <p style="color: #ffffff; margin-bottom: 10px;"><strong>E-Mail:</strong> {comment_data['email']}</p>
                <p style="color: #ffffff; margin-bottom: 20px;"><strong>Kommentar:</strong></p>
                <div style="background-color: #002147; padding: 15px; border-radius: 5px; border-left: 3px solid #D4AF37;">
                    <p style="color: #ffffff; margin: 0;">{comment_data['content']}</p>
                </div>
                
                <p style="color: #888888; margin-top: 30px; font-size: 12px;">
                    Dieser Kommentar wartet auf Freigabe im Admin-Panel.
                </p>
                
                <a href="https://invest.euroadria.me/admin" 
                   style="display: inline-block; margin-top: 20px; padding: 12px 24px; background-color: #D4AF37; color: #002147; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    Zum Admin-Panel
                </a>
            </div>
        </body>
        </html>
        """
        
        resend.Emails.send({
            "from": "EuroAdria Corporate Solutions <noreply@euroadria.me>",
            "to": [NOTIFICATION_EMAIL],
            "subject": f"Neuer Kommentar auf EuroAdria: {article_title[:50]}",
            "html": html_content
        })
        
        logger.info(f"Notification email sent for comment by {comment_data['name']}")
    except Exception as e:
        logger.error(f"Failed to send notification email: {e}")


# Contact form email function using Resend
async def send_contact_email(contact_data: dict):
    """Send email notification for contact form submission using Resend"""
    if not RESEND_API_KEY:
        logger.warning("Resend API key not configured - skipping contact email")
        return False
    
    try:
        resend.api_key = RESEND_API_KEY
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; background-color: #0A1628; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #1a2744; border-radius: 10px; padding: 30px;">
                <h2 style="color: #c8a96a; margin-bottom: 20px;">Neue Kontaktanfrage</h2>
                <div style="background-color: #0A1628; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <p style="color: #ffffff; margin-bottom: 10px;"><strong>Name:</strong> {contact_data['name']}</p>
                    <p style="color: #ffffff; margin-bottom: 10px;"><strong>E-Mail:</strong> {contact_data['email']}</p>
                    <p style="color: #ffffff; margin-bottom: 10px;"><strong>Telefon:</strong> {contact_data.get('phone', 'Nicht angegeben')}</p>
                    <p style="color: #ffffff; margin-bottom: 10px;"><strong>Betreff:</strong> {contact_data['subject']}</p>
                </div>
                <div style="background-color: #0A1628; padding: 20px; border-radius: 8px;">
                    <p style="color: #c8a96a; margin-bottom: 10px;"><strong>Nachricht:</strong></p>
                    <p style="color: #ffffff; white-space: pre-wrap;">{contact_data['message']}</p>
                </div>
                <hr style="border-color: #c8a96a; margin: 20px 0;">
                <p style="color: #888; font-size: 12px;">Diese Nachricht wurde über das Kontaktformular auf euroadria.me gesendet.</p>
            </div>
        </body>
        </html>
        """
        
        params = {
            "from": "EuroAdria Kontakt <onboarding@resend.dev>",
            "to": [NOTIFICATION_EMAIL],
            "subject": f"Neue Kontaktanfrage: {contact_data['subject']}",
            "html": html_content,
            "reply_to": contact_data['email']
        }
        
        email = resend.Emails.send(params)
        logger.info(f"Contact email sent via Resend from {contact_data['name']}, id: {email.get('id')}")
        return True
    except Exception as e:
        logger.error(f"Failed to send contact email: {e}")
        return False


# Contact form model
class ContactForm(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    subject: str
    message: str

# Lead form model (for Exposé downloads)
class LeadForm(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    source: str  # e.g. "budva_expose", "niksic_expose"
    expose_name: Optional[str] = None


# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}


# =============================================
# CONTACT FORM ENDPOINT
# =============================================

@api_router.post("/contact")
async def submit_contact_form(form: ContactForm):
    """Handle contact form submissions"""
    contact_dict = form.model_dump()
    contact_dict["submitted_at"] = datetime.now(timezone.utc).isoformat()
    contact_dict["status"] = "new"
    
    # Store in database
    await db.contact_submissions.insert_one(contact_dict)
    
    # Also count as lead
    await db.leads.insert_one({
        "name": contact_dict.get("name", ""),
        "email": contact_dict.get("email", ""),
        "source": "kontaktformular",
        "expose_name": f"Kontakt: {contact_dict.get('subject', 'Allgemein')}",
        "type": "contact",
        "submitted_at": contact_dict["submitted_at"]
    })
    
    # Auto-create CRM lead + deal
    existing_crm = await db.crm_leads.find_one({"email": contact_dict.get("email", "")})
    if existing_crm:
        # Same email exists — add a new deal to the existing lead
        await db.crm_deals.insert_one({
            "id": str(uuid.uuid4())[:8],
            "lead_id": existing_crm["id"],
            "stage": "new_lead",
            "deal_value": 0,
            "probability": 10,
            "expected_revenue": 0,
            "assigned_to": None,
            "notes": f"Betreff: {contact_dict.get('subject', 'Allgemein')}",
            "created_at": contact_dict["submitted_at"],
            "updated_at": contact_dict["submitted_at"]
        })
        # Reactivate lead if it was lost
        if existing_crm.get("status") == "lost":
            await db.crm_leads.update_one({"id": existing_crm["id"]}, {"$set": {"status": "new"}})
    else:
        crm_lead_id = str(uuid.uuid4())[:8]
        await db.crm_leads.insert_one({
            "id": crm_lead_id,
            "name": contact_dict.get("name", ""),
            "email": contact_dict.get("email", ""),
            "phone": contact_dict.get("phone"),
            "lead_source": "kontaktformular",
            "utm_source": None, "utm_medium": None, "utm_campaign": None,
            "entry_page": None,
            "tool_used": "contact_form",
            "status": "new",
            "created_at": contact_dict["submitted_at"]
        })
        await db.crm_deals.insert_one({
            "id": str(uuid.uuid4())[:8],
            "lead_id": crm_lead_id,
            "stage": "new_lead",
            "deal_value": 0,
            "probability": 10,
            "expected_revenue": 0,
            "assigned_to": None,
            "notes": f"Betreff: {contact_dict.get('subject', 'Allgemein')}",
            "created_at": contact_dict["submitted_at"],
            "updated_at": contact_dict["submitted_at"]
        })
    
    # Send email notification
    email_sent = await send_contact_email(contact_dict)
    
    return {
        "success": True,
        "message": "Vielen Dank für Ihre Nachricht! Wir melden uns zeitnah bei Ihnen.",
        "email_sent": email_sent
    }


# =============================================
# LEAD CAPTURE ENDPOINT (Exposé Downloads)
# =============================================

@api_router.post("/leads")
async def capture_lead(lead: LeadForm):
    """Capture lead data before Exposé download"""
    lead_dict = lead.model_dump()
    lead_dict["submitted_at"] = datetime.now(timezone.utc).isoformat()
    lead_dict["type"] = "expose_download"
    
    # Store in database
    await db.leads.insert_one(lead_dict)
    
    # Auto-create CRM lead + deal
    existing_crm = await db.crm_leads.find_one({"email": lead_dict.get("email", "")})
    if existing_crm:
        await db.crm_deals.insert_one({
            "id": str(uuid.uuid4())[:8],
            "lead_id": existing_crm["id"],
            "stage": "new_lead",
            "deal_value": 0, "probability": 10, "expected_revenue": 0,
            "assigned_to": None,
            "notes": f"Exposé: {lead_dict.get('expose_name', lead_dict.get('source', ''))}",
            "created_at": lead_dict["submitted_at"],
            "updated_at": lead_dict["submitted_at"]
        })
        if existing_crm.get("status") == "lost":
            await db.crm_leads.update_one({"id": existing_crm["id"]}, {"$set": {"status": "new"}})
    else:
        crm_lead_id = str(uuid.uuid4())[:8]
        await db.crm_leads.insert_one({
            "id": crm_lead_id,
            "name": lead_dict.get("name", ""),
            "email": lead_dict.get("email", ""),
            "phone": lead_dict.get("phone"),
            "lead_source": "expose_download",
            "utm_source": None, "utm_medium": None, "utm_campaign": None,
            "entry_page": None,
            "tool_used": lead_dict.get("expose_name", lead_dict.get("source", "expose")),
            "status": "new",
            "created_at": lead_dict["submitted_at"]
        })
        await db.crm_deals.insert_one({
            "id": str(uuid.uuid4())[:8],
            "lead_id": crm_lead_id,
            "stage": "new_lead",
            "deal_value": 0, "probability": 10, "expected_revenue": 0,
            "assigned_to": None,
            "notes": f"Exposé: {lead_dict.get('expose_name', lead_dict.get('source', ''))}",
            "created_at": lead_dict["submitted_at"],
            "updated_at": lead_dict["submitted_at"]
        })
    
    # Send email notification
    email_sent = False
    if RESEND_API_KEY:
        try:
            resend.api_key = RESEND_API_KEY
            html_content = f"""
            <html>
            <body style="font-family: Arial, sans-serif; background-color: #0A1628; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: #1a2744; border-radius: 10px; padding: 30px;">
                    <h2 style="color: #c8a96a; margin-bottom: 20px;">Neuer Exposé-Download Lead</h2>
                    <div style="background-color: #0A1628; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                        <p style="color: #ffffff; margin-bottom: 10px;"><strong>Name:</strong> {lead_dict['name']}</p>
                        <p style="color: #ffffff; margin-bottom: 10px;"><strong>E-Mail:</strong> {lead_dict['email']}</p>
                        <p style="color: #ffffff; margin-bottom: 10px;"><strong>Telefon:</strong> {lead_dict.get('phone', 'Nicht angegeben')}</p>
                        <p style="color: #ffffff; margin-bottom: 10px;"><strong>Exposé:</strong> {lead_dict.get('expose_name', lead_dict['source'])}</p>
                    </div>
                    <hr style="border-color: #c8a96a; margin: 20px 0;">
                    <p style="color: #888; font-size: 12px;">Dieser Lead wurde über den Exposé-Download auf invest.euroadria.me generiert.</p>
                </div>
            </body>
            </html>
            """
            params = {
                "from": "EuroAdria Leads <onboarding@resend.dev>",
                "to": [NOTIFICATION_EMAIL],
                "subject": f"Neuer Lead: {lead_dict.get('expose_name', lead_dict['source'])} - {lead_dict['name']}",
                "html": html_content,
                "reply_to": lead_dict['email']
            }
            resend.Emails.send(params)
            email_sent = True
        except Exception as e:
            logger.error(f"Failed to send lead email: {e}")
    
    return {"success": True, "email_sent": email_sent}

@api_router.get("/admin/leads")
async def get_leads(admin: str = Depends(verify_admin)):
    """Get all collected leads (Admin only)"""
    leads = await db.leads.find({}, {"_id": 0}).sort("submitted_at", -1).to_list(500)
    return leads


# =============================================
# ANALYTICS & TRACKING ENDPOINTS
# =============================================

class PageViewEvent(BaseModel):
    path: str
    referrer: Optional[str] = None
    user_agent: Optional[str] = None
    utm_source: Optional[str] = None
    utm_medium: Optional[str] = None
    utm_campaign: Optional[str] = None

@api_router.post("/track/pageview")
async def track_pageview(event: PageViewEvent):
    """Track a page view (called from frontend)"""
    doc = {
        "path": event.path,
        "referrer": event.referrer or "",
        "device": parse_device_type(event.user_agent or ""),
        "utm_source": event.utm_source or "",
        "utm_medium": event.utm_medium or "",
        "utm_campaign": event.utm_campaign or "",
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.page_views.insert_one(doc)
    return {"ok": True}

@api_router.post("/track/calculator")
async def track_calculator_usage():
    """Track ROI calculator usage"""
    doc = {"timestamp": datetime.now(timezone.utc).isoformat(), "type": "roi_calculator"}
    await db.calculator_usage.insert_one(doc)
    return {"ok": True}

@api_router.get("/admin/analytics/overview")
async def get_analytics_overview(days: int = 30, admin: str = Depends(verify_admin)):
    """Get analytics overview data for dashboard"""
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    # Total page views in period
    total_views = await db.page_views.count_documents({"timestamp": {"$gte": cutoff}})
    
    # Views per day (for chart)
    pipeline_daily = [
        {"$match": {"timestamp": {"$gte": cutoff}}},
        {"$addFields": {"date": {"$substr": ["$timestamp", 0, 10]}}},
        {"$group": {"_id": "$date", "views": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    daily_views = await db.page_views.aggregate(pipeline_daily).to_list(60)
    
    # Top pages
    pipeline_pages = [
        {"$match": {"timestamp": {"$gte": cutoff}}},
        {"$group": {"_id": "$path", "views": {"$sum": 1}}},
        {"$sort": {"views": -1}},
        {"$limit": 10}
    ]
    top_pages = await db.page_views.aggregate(pipeline_pages).to_list(10)
    
    # Device breakdown
    pipeline_devices = [
        {"$match": {"timestamp": {"$gte": cutoff}}},
        {"$group": {"_id": "$device", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    devices = await db.page_views.aggregate(pipeline_devices).to_list(5)
    
    # Referrer breakdown (traffic sources) - combined with UTM data
    pipeline_referrers = [
        {"$match": {"timestamp": {"$gte": cutoff}}},
        {"$addFields": {
            "has_utm": {"$and": [
                {"$ne": [{"$ifNull": ["$utm_source", ""]}, ""]},
                {"$ne": [{"$ifNull": ["$utm_source", ""]}, None]}
            ]},
            "has_ref": {"$and": [
                {"$ne": [{"$ifNull": ["$referrer", ""]}, ""]},
                {"$ne": [{"$ifNull": ["$referrer", ""]}, None]}
            ]}
        }},
        {"$addFields": {
            "source": {
                "$switch": {
                    "branches": [
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "tiktok", "options": "i"}}]}, "then": "TikTok"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "^instagram$|^ig$|^insta$|^lg$", "options": "i"}}]}, "then": "Instagram"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "facebook|^fb$", "options": "i"}}]}, "then": "Facebook"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "youtube|^yt$", "options": "i"}}]}, "then": "YouTube"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "linkedin", "options": "i"}}]}, "then": "LinkedIn"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "google", "options": "i"}}]}, "then": "Google"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "^tt$|^tik$", "options": "i"}}]}, "then": "TikTok"},
                        {"case": {"$and": ["$has_utm", {"$regexMatch": {"input": {"$ifNull": ["$utm_source", ""]}, "regex": "^twitter$|^x$|x\\.com", "options": "i"}}]}, "then": "Twitter/X"},
                        {"case": "$has_utm", "then": "Andere (UTM)"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "google"}}]}, "then": "Google"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "linkedin"}}]}, "then": "LinkedIn"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "facebook|fb.com"}}]}, "then": "Facebook"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "instagram|l\\.instagram"}}]}, "then": "Instagram"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "twitter|x.com"}}]}, "then": "Twitter/X"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "tiktok"}}]}, "then": "TikTok"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "youtube"}}]}, "then": "YouTube"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "euroadria"}}]}, "then": "EuroAdria.me"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "rtl\\.de|rtl\\.com|n-tv|ntv"}}]}, "then": "RTL / n-tv"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "focus\\.de|focus\\.com"}}]}, "then": "Focus"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "bild\\.de"}}]}, "then": "Bild"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "spiegel\\.de"}}]}, "then": "Spiegel"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "welt\\.de"}}]}, "then": "Welt"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "faz\\.net"}}]}, "then": "FAZ"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "handelsblatt\\.com"}}]}, "then": "Handelsblatt"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "finanzen\\.net|finanzen\\.de"}}]}, "then": "Finanzen.net"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "t-online\\.de"}}]}, "then": "t-online"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "reddit\\.com"}}]}, "then": "Reddit"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "whatsapp\\.com|wa\\.me"}}]}, "then": "WhatsApp"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "t\\.me|telegram"}}]}, "then": "Telegram"},
                        {"case": {"$and": ["$has_ref", {"$regexMatch": {"input": "$referrer", "regex": "bing\\.com"}}]}, "then": "Bing"},
                        {"case": "$has_ref", "then": "Andere"},
                    ],
                    "default": "Direkt"
                }
            }
        }},
        {"$group": {"_id": "$source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    referrers = await db.page_views.aggregate(pipeline_referrers).to_list(10)
    
    # Total leads in period
    total_leads = await db.leads.count_documents({"submitted_at": {"$gte": cutoff}})
    
    # Leads by source
    pipeline_lead_sources = [
        {"$match": {"submitted_at": {"$gte": cutoff}}},
        {"$group": {"_id": "$source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    lead_sources = await db.leads.aggregate(pipeline_lead_sources).to_list(10)
    
    # Calculator usage count
    calc_usage = await db.calculator_usage.count_documents({"timestamp": {"$gte": cutoff}})
    
    # Contact form submissions count
    total_contacts = await db.contact_submissions.count_documents({"submitted_at": {"$gte": cutoff}})
    
    # Recent leads
    recent_leads = await db.leads.find({}, {"_id": 0}).sort("submitted_at", -1).to_list(20)
    
    # Conversion rate
    conversion_rate = round((total_leads / total_views * 100), 2) if total_views > 0 else 0
    
    # UTM Campaign / Source tracking (normalized)
    pipeline_utm = [
        {"$match": {"timestamp": {"$gte": cutoff}, "utm_source": {"$nin": ["", None]}}},
        {"$addFields": {
            "norm_source": {
                "$switch": {
                    "branches": [
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "tiktok|^tt$|^tik$", "options": "i"}}, "then": "TikTok"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "^instagram$|^ig$|^insta$|^lg$", "options": "i"}}, "then": "Instagram"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "facebook|^fb$", "options": "i"}}, "then": "Facebook"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "youtube|^yt$", "options": "i"}}, "then": "YouTube"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "linkedin", "options": "i"}}, "then": "LinkedIn"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "google", "options": "i"}}, "then": "Google"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "^twitter$|^x$|x\\.com", "options": "i"}}, "then": "Twitter/X"},
                    ],
                    "default": "$utm_source"
                }
            }
        }},
        {"$group": {
            "_id": {
                "source": "$norm_source",
                "medium": "$utm_medium",
                "campaign": "$utm_campaign"
            },
            "count": {"$sum": 1}
        }},
        {"$sort": {"count": -1}},
        {"$limit": 15}
    ]
    utm_data = await db.page_views.aggregate(pipeline_utm).to_list(15)
    
    # UTM sources summary (grouped by source only, normalized)
    pipeline_utm_sources = [
        {"$match": {"timestamp": {"$gte": cutoff}, "utm_source": {"$nin": ["", None]}}},
        {"$addFields": {
            "norm_source": {
                "$switch": {
                    "branches": [
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "tiktok|^tt$|^tik$", "options": "i"}}, "then": "TikTok"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "^instagram$|^ig$|^insta$|^lg$", "options": "i"}}, "then": "Instagram"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "facebook|^fb$", "options": "i"}}, "then": "Facebook"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "youtube|^yt$", "options": "i"}}, "then": "YouTube"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "linkedin", "options": "i"}}, "then": "LinkedIn"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "google", "options": "i"}}, "then": "Google"},
                        {"case": {"$regexMatch": {"input": "$utm_source", "regex": "^twitter$|^x$|x\\.com", "options": "i"}}, "then": "Twitter/X"},
                    ],
                    "default": "$utm_source"
                }
            }
        }},
        {"$group": {"_id": "$norm_source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    utm_sources = await db.page_views.aggregate(pipeline_utm_sources).to_list(10)
    
    return {
        "total_views": total_views,
        "total_leads": total_leads,
        "total_contacts": total_contacts,
        "calculator_usage": calc_usage,
        "conversion_rate": conversion_rate,
        "daily_views": [{"date": d["_id"], "views": d["views"]} for d in daily_views],
        "top_pages": [{"path": p["_id"], "views": p["views"]} for p in top_pages],
        "devices": [{"device": d["_id"], "count": d["count"]} for d in devices],
        "referrers": [{"source": r["_id"], "count": r["count"]} for r in referrers],
        "lead_sources": [{"source": l["_id"], "count": l["count"]} for l in lead_sources],
        "recent_leads": recent_leads,
        "utm_sources": [{"source": u["_id"], "count": u["count"]} for u in utm_sources],
        "utm_campaigns": [{"source": u["_id"].get("source", "-"), "medium": u["_id"].get("medium", "") or "-", "campaign": u["_id"].get("campaign", "") or "-", "count": u["count"]} for u in utm_data]
    }


@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks


# =============================================
# NEWSLETTER ENDPOINTS (Brevo)
# =============================================


class NewsletterSubscribe(BaseModel):
    email: EmailStr
    name: Optional[str] = None

def brevo_request(method: str, endpoint: str, data: dict = None):
    """Make a request to Brevo API"""
    url = f"https://api.brevo.com/v3/{endpoint}"
    headers = {"api-key": BREVO_API_KEY, "content-type": "application/json"}
    if method == "GET":
        r = http_requests.get(url, headers=headers, params=data)
    elif method == "POST":
        r = http_requests.post(url, headers=headers, json=data)
    elif method == "PUT":
        r = http_requests.put(url, headers=headers, json=data)
    else:
        raise ValueError(f"Unsupported method: {method}")
    return r

@api_router.post("/newsletter/subscribe")
async def newsletter_subscribe(sub: NewsletterSubscribe):
    """Subscribe to newsletter via Brevo"""
    if not BREVO_API_KEY:
        raise HTTPException(status_code=500, detail="Newsletter nicht konfiguriert")
    
    # Create/update contact in Brevo with list assignment
    payload = {
        "email": sub.email,
        "updateEnabled": True,
        "listIds": [BREVO_LIST_ID],
    }
    if sub.name:
        payload["attributes"] = {"VORNAME": sub.name}
    
    r = brevo_request("POST", "contacts", payload)
    
    if r.status_code in (200, 201, 204):
        # Check if already subscribed locally (to avoid duplicate welcome emails)
        existing = await db.newsletter_subscribers.find_one({"email": sub.email})
        is_new = existing is None
        
        # Save/update locally for analytics
        await db.newsletter_subscribers.update_one(
            {"email": sub.email},
            {"$set": {"email": sub.email, "name": sub.name or "", "subscribed_at": datetime.now(timezone.utc).isoformat(), "active": True}},
            upsert=True
        )
        
        # Also count as lead
        if is_new:
            await db.leads.insert_one({
                "name": sub.name or "",
                "email": sub.email,
                "source": "newsletter",
                "expose_name": "Newsletter-Anmeldung",
                "type": "newsletter",
                "submitted_at": datetime.now(timezone.utc).isoformat()
            })
        
        # Send welcome email ONLY for new subscribers
        if is_new:
            try:
                welcome_html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px; margin: 0;">
                <div style="max-width: 640px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <div style="background: #ffffff; padding: 28px 32px; text-align: center; border-bottom: 2px solid #C8A96A;">
                        <img src="https://invest.euroadria.me/euroadria-logo.png" alt="EuroAdria" style="height: 50px;">
                    </div>
                    <div style="padding: 32px; color: #333; font-size: 15px; line-height: 1.7;">
                        <h2 style="color: #04151F; font-size: 22px; margin-bottom: 16px;">Willkommen bei EuroAdria Corporate Solutions!</h2>
                        <p>Hallo{(' ' + sub.name) if sub.name else ''},</p>
                        <p>vielen Dank für Ihre Anmeldung zum EuroAdria Newsletter! 
                        Sie erhalten ab sofort exklusive Investment-Insights, Marktanalysen 
                        und Neuigkeiten zu Immobilien-Projekten am Balkan.</p>
                        <p style="margin-top: 20px;">
                            <a href="https://invest.euroadria.me" style="display: inline-block; background: #C8A96A; color: #04151F; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: bold;">Zur Plattform</a>
                        </p>
                    </div>
                    <div style="padding: 16px 32px; background: #f8f8f8; border-top: 1px solid #eee; border-bottom: 1px solid #eee;">
                        <p style="color: #999; font-size: 12px; margin: 0; text-align: center;">
                            Sie erhalten diese E-Mail, weil Sie sich für den EuroAdria Newsletter angemeldet haben.<br>
                            Sie können sich jederzeit mit einem Klick 
                            <a href="https://invest.euroadria.me/newsletter/abmelden?email={sub.email}" style="color: #C8A96A;">hier abmelden</a> 
                            , unkompliziert und sofort wirksam.
                        </p>
                    </div>
                    <div style="padding: 24px 32px; border-top: 1px solid #e5e7eb; background: #fafafa;">
                        <table style="width: 100%;">
                            <tr>
                                <td style="vertical-align: top; padding-right: 20px; width: 130px;">
                                    <img src="https://invest.euroadria.me/euroadria-logo.png" alt="EuroAdria" style="width: 110px;">
                                </td>
                                <td style="vertical-align: top; font-size: 12px; color: #555; line-height: 1.6;">
                                    <p style="margin: 0 0 2px; font-size: 14px; font-weight: bold; color: #04151F;">EuroAdria Corporate Solutions</p>
                                    <p style="margin: 0 0 8px; font-size: 11px; color: #888;">a brand of <strong style="color: #333;">Montaris &amp; Co. d.o.o.</strong></p>
                                    <p style="margin: 0 0 8px; color: #555;">Montaris &amp; Co. d.o.o.<br>Marka Miljanova 12<br>21000 Novi Sad, Serbia</p>
                                    <p style="margin: 0 0 8px;">
                                        Tel.: <a href="tel:+38268559776" style="color: #C8A96A; text-decoration: none;">+382 68 559 776</a><br>
                                        E-Mail: <a href="mailto:office@euroadria.me" style="color: #C8A96A; text-decoration: none;">office@euroadria.me</a><br>
                                        Web: <a href="https://www.euroadria.me" style="color: #C8A96A; text-decoration: none;">www.euroadria.me</a><br>
                                        Investment: <a href="https://invest.euroadria.me" style="color: #C8A96A; text-decoration: none;">invest.euroadria.me</a>
                                    </p>
                                    <p style="margin: 0; font-size: 10px; color: #999; line-height: 1.5;">
                                        Company registration no.: 22147382 | Tax ID (PIB): 115356237<br>
                                        Director: Milena Bubanja
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </div>
                </div>
            </body>
            </html>
            """
                brevo_request("POST", "smtp/email", {
                    "to": [{"email": sub.email, "name": sub.name or sub.email}],
                    "sender": {"email": "office@euroadria.me", "name": "EuroAdria"},
                    "subject": "Willkommen beim EuroAdria Newsletter!",
                    "htmlContent": welcome_html
                })
            except Exception as e:
                logger.error(f"Welcome email failed: {e}")
        
        return {"success": True, "message": "Erfolgreich zum Newsletter angemeldet!" if is_new else "Sie sind bereits angemeldet!"}
    
    elif r.status_code == 400 and "already exist" in r.text.lower():
        return {"success": True, "message": "Sie sind bereits angemeldet!"}
    else:
        logger.error(f"Brevo subscribe error: {r.status_code} {r.text}")
        raise HTTPException(status_code=400, detail="Anmeldung fehlgeschlagen. Bitte versuchen Sie es erneut.")

@api_router.post("/newsletter/unsubscribe")
async def newsletter_unsubscribe(data: dict):
    """Unsubscribe from newsletter"""
    email = data.get("email", "")
    if not email:
        raise HTTPException(status_code=400, detail="E-Mail erforderlich")
    
    # Remove from Brevo list
    if BREVO_API_KEY:
        try:
            brevo_request("POST", f"contacts/lists/{BREVO_LIST_ID}/contacts/remove", {"emails": [email]})
        except Exception as e:
            logger.error(f"Brevo unsubscribe error: {e}")
    
    # Mark as inactive locally
    await db.newsletter_subscribers.update_one(
        {"email": email},
        {"$set": {"active": False, "unsubscribed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "message": "Sie wurden erfolgreich abgemeldet."}

@api_router.get("/admin/newsletter/subscribers")
async def get_newsletter_subscribers(admin: str = Depends(verify_admin)):
    """Get all newsletter subscribers"""
    # Get from Brevo
    r = brevo_request("GET", f"contacts/lists/{BREVO_LIST_ID}/contacts?limit=500")
    brevo_contacts = []
    if r.ok:
        data = r.json()
        brevo_contacts = [{"email": c["email"], "name": c.get("attributes", {}).get("VORNAME", ""), "subscribed": True} for c in data.get("contacts", [])]
    
    # Get local count
    local_count = await db.newsletter_subscribers.count_documents({"active": True})
    
    return {
        "subscribers": brevo_contacts,
        "total": len(brevo_contacts),
        "local_count": local_count
    }


@api_router.delete("/admin/newsletter/subscribers/{email}")
async def admin_delete_subscriber(email: str, admin: str = Depends(verify_admin)):
    """Admin: Delete a newsletter subscriber"""
    from urllib.parse import unquote
    email = unquote(email)
    
    # Remove from Brevo
    if BREVO_API_KEY:
        try:
            brevo_request("POST", f"contacts/lists/{BREVO_LIST_ID}/contacts/remove", {"emails": [email]})
        except Exception as e:
            logger.error(f"Brevo remove error: {e}")
    
    # Remove from local DB
    await db.newsletter_subscribers.delete_one({"email": email})
    
    return {"success": True, "message": f"{email} wurde gelöscht."}


@api_router.post("/admin/newsletter/send")
async def send_newsletter(data: dict, admin: str = Depends(verify_admin)):
    """Send a newsletter campaign via Brevo"""
    if not BREVO_API_KEY:
        raise HTTPException(status_code=500, detail="Brevo nicht konfiguriert")
    
    subject = data.get("subject", "")
    content = data.get("content", "")
    
    if not subject or not content:
        raise HTTPException(status_code=400, detail="Betreff und Inhalt sind erforderlich")
    
    # Wrap content in branded template
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px; margin: 0;">
        <div style="max-width: 640px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
            <div style="background: #ffffff; padding: 28px 32px; text-align: center; border-bottom: 2px solid #C8A96A;">
                <img src="https://invest.euroadria.me/euroadria-logo.png" alt="EuroAdria" style="height: 50px;">
            </div>
            <div style="padding: 32px; color: #333; font-size: 15px; line-height: 1.7;">
                {content}
            </div>
            <div style="padding: 24px 32px; border-top: 1px solid #e5e7eb; background: #fafafa;">
                <table style="width: 100%;">
                    <tr>
                        <td style="vertical-align: top; padding-right: 20px; width: 130px;">
                            <img src="https://invest.euroadria.me/euroadria-logo.png" alt="EuroAdria" style="width: 110px;">
                        </td>
                        <td style="vertical-align: top; font-size: 12px; color: #555; line-height: 1.6;">
                            <p style="margin: 0 0 2px; font-size: 14px; font-weight: bold; color: #04151F;">EuroAdria Corporate Solutions</p>
                            <p style="margin: 0 0 8px; font-size: 11px; color: #888;">a brand of <strong style="color: #333;">Montaris &amp; Co. d.o.o.</strong></p>
                            <p style="margin: 0 0 8px; color: #555;">Montaris &amp; Co. d.o.o.<br>Marka Miljanova 12<br>21000 Novi Sad, Serbia</p>
                            <p style="margin: 0 0 8px;">
                                Tel.: <a href="tel:+38268559776" style="color: #C8A96A; text-decoration: none;">+382 68 559 776</a><br>
                                E-Mail: <a href="mailto:office@euroadria.me" style="color: #C8A96A; text-decoration: none;">office@euroadria.me</a><br>
                                Web: <a href="https://www.euroadria.me" style="color: #C8A96A; text-decoration: none;">www.euroadria.me</a><br>
                                Investment: <a href="https://invest.euroadria.me" style="color: #C8A96A; text-decoration: none;">invest.euroadria.me</a>
                            </p>
                            <p style="margin: 0; font-size: 10px; color: #999; line-height: 1.5;">
                                Company registration no.: 22147382 | Tax ID (PIB): 115356237<br>
                                Director: Milena Bubanja
                            </p>
                        </td>
                    </tr>
                </table>
            </div>
            <div style="background: #04151F; padding: 16px 32px; text-align: center;">
                <p style="color: #888; font-size: 11px; margin: 0;">
                    <a href="{{{{unsubscribe}}}}" style="color: #C8A96A;">Newsletter abbestellen</a>
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Create campaign in Brevo
    campaign_data = {
        "name": f"Newsletter: {subject} ({datetime.now(timezone.utc).strftime('%d.%m.%Y')})",
        "sender": {"email": "office@euroadria.me", "name": "EuroAdria"},
        "subject": subject,
        "htmlContent": html_body,
        "recipients": {"listIds": [BREVO_LIST_ID]},
        "type": "classic"
    }
    
    r = brevo_request("POST", "emailCampaigns", campaign_data)
    
    if not r.ok:
        logger.error(f"Brevo campaign create error: {r.status_code} {r.text}")
        raise HTTPException(status_code=400, detail=f"Kampagne konnte nicht erstellt werden: {r.text}")
    
    campaign_id = r.json().get("id")
    
    # Send immediately
    r2 = brevo_request("POST", f"emailCampaigns/{campaign_id}/sendNow", {})
    
    if r2.status_code in (200, 204):
        # Log in DB
        await db.newsletter_campaigns.insert_one({
            "campaign_id": campaign_id,
            "subject": subject,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "list_id": BREVO_LIST_ID
        })
        return {"success": True, "campaign_id": campaign_id, "message": "Newsletter wurde versendet!"}
    else:
        logger.error(f"Brevo send error: {r2.status_code} {r2.text}")
        raise HTTPException(status_code=400, detail=f"Newsletter konnte nicht gesendet werden: {r2.text}")


# =============================================
# ARTICLE ENDPOINTS (Public)
# =============================================

# Lightweight article fields for blog list (excludes heavy content)
BLOG_LIST_FIELDS = {
    "_id": 0,
    "id": 1,
    "slug": 1,
    "title": 1,
    "excerpt": 1,
    "image": 1,
    "category": 1,
    "cluster": 1,
    "date": 1,
    "readTime": 1,
    "featured": 1,
    "author": 1
}

@api_router.get("/articles/list")
async def get_articles_list(
    cluster: Optional[str] = None,
    category: Optional[str] = None,
    page: int = 1,
    limit: int = 12,
    search: Optional[str] = None,
    lang: Optional[str] = None
):
    """Get paginated article list for blog page (lightweight - no content)"""
    query = {}
    if cluster and cluster != "All":
        query["category"] = cluster
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"excerpt": {"$regex": search, "$options": "i"}}
        ]
    
    # Get total count
    total = await db.articles.count_documents(query)
    
    # Get paginated results
    skip = (page - 1) * limit
    articles = await db.articles.find(query, BLOG_LIST_FIELDS).sort([("sortOrder", 1), ("date", -1)]).skip(skip).limit(limit).to_list(limit)
    
    # Translate titles and excerpts if lang=en
    if lang == "en":
        for art in articles:
            if art.get("title"):
                art["title"] = await _get_or_translate(art["title"], "de", "en")
            if art.get("excerpt"):
                art["excerpt"] = await _get_or_translate(art["excerpt"], "de", "en")
            if art.get("category"):
                art["category"] = await _get_or_translate(art["category"], "de", "en")
    
    return {
        "articles": articles,
        "total": total,
        "page": page,
        "limit": limit,
        "totalPages": (total + limit - 1) // limit,
        "hasMore": skip + len(articles) < total
    }


@api_router.get("/articles", response_model=List[Article])
async def get_articles(
    cluster: Optional[str] = None,
    featured: Optional[bool] = None,
    category: Optional[str] = None,
    limit: Optional[int] = None
):
    """Get all articles with optional filters (full data)"""
    query = {}
    if cluster:
        query["cluster"] = cluster
    if featured is not None:
        query["featured"] = featured
    if category:
        query["category"] = category
    
    cursor = db.articles.find(query, {"_id": 0}).sort([("sortOrder", 1), ("date", -1)])
    if limit:
        cursor = cursor.limit(limit)
    
    articles = await cursor.to_list(1000)
    return articles


@api_router.get("/articles/{slug}", response_model=Article)
async def get_article_by_slug(slug: str):
    """Get a single article by slug"""
    article = await db.articles.find_one({"slug": slug}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    return article


@api_router.get("/og/blog/{slug}")
async def get_article_og_html(slug: str):
    """Serve pre-rendered HTML with OG tags for social media crawlers"""
    article = await db.articles.find_one({"slug": slug}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    
    site_url = "https://invest.euroadria.me"
    title = f"{article.get('title', '')} | EuroAdria Corporate Solutions"
    description = article.get('excerpt', '')[:200]
    image = article.get('image', f"{site_url}/euroadria-logo.png")
    url = f"{site_url}/blog/{slug}"
    
    html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<meta name="description" content="{description}"/>
<meta property="og:type" content="article"/>
<meta property="og:url" content="{url}"/>
<meta property="og:title" content="{title}"/>
<meta property="og:description" content="{description}"/>
<meta property="og:image" content="{image}"/>
<meta property="og:image:width" content="1200"/>
<meta property="og:image:height" content="630"/>
<meta property="og:locale" content="de_DE"/>
<meta property="og:site_name" content="EuroAdria Corporate Solutions"/>
<meta name="twitter:card" content="summary_large_image"/>
<meta name="twitter:title" content="{title}"/>
<meta name="twitter:description" content="{description}"/>
<meta name="twitter:image" content="{image}"/>
<meta http-equiv="refresh" content="0;url={url}"/>
</head>
<body><p>Weiterleitung zu <a href="{url}">{title}</a></p></body>
</html>"""
    return Response(content=html, media_type="text/html")


@api_router.get("/articles/id/{article_id}", response_model=Article)
async def get_article_by_id(article_id: int):
    """Get a single article by ID"""
    article = await db.articles.find_one({"id": article_id}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    return article


@api_router.get("/clusters")
async def get_clusters():
    """Get all unique categories with article counts"""
    pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    clusters = await db.articles.aggregate(pipeline).to_list(100)
    return [{"id": c["_id"], "name": c["_id"], "count": c["count"]} for c in clusters if c["_id"]]


@api_router.get("/categories")
async def get_categories():
    """Get all unique categories with article counts"""
    pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    categories = await db.articles.aggregate(pipeline).to_list(100)
    return [{"category": c["_id"], "count": c["count"]} for c in categories]


# =============================================
# ADMIN ENDPOINTS (Protected)
# =============================================

@api_router.post("/admin/articles", response_model=Article)
async def create_article(article: ArticleCreate, admin: str = Depends(verify_admin)):
    """Create a new article (Admin only)"""
    # Check if slug already exists
    existing = await db.articles.find_one({"slug": article.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Article with this slug already exists")
    
    # Generate new ID (find max ID and increment)
    max_id_doc = await db.articles.find_one(sort=[("id", -1)])
    new_id = (max_id_doc["id"] + 1) if max_id_doc else 101
    
    article_dict = article.model_dump()
    article_dict["id"] = new_id
    
    await db.articles.insert_one(article_dict)
    
    # Return without _id
    return article_dict


@api_router.put("/admin/articles/{article_id}", response_model=Article)
async def update_article(article_id: int, article_update: ArticleUpdate, admin: str = Depends(verify_admin)):
    """Update an article (Admin only)"""
    # Check if article exists
    existing = await db.articles.find_one({"id": article_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Article not found")
    
    # Build update dict excluding None values
    update_data = {k: v for k, v in article_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # If slug is being updated, check it doesn't conflict
    if "slug" in update_data and update_data["slug"] != existing.get("slug"):
        slug_exists = await db.articles.find_one({"slug": update_data["slug"], "id": {"$ne": article_id}})
        if slug_exists:
            raise HTTPException(status_code=400, detail="Article with this slug already exists")
    
    await db.articles.update_one({"id": article_id}, {"$set": update_data})
    
    updated = await db.articles.find_one({"id": article_id}, {"_id": 0})
    return updated


@api_router.delete("/admin/articles/{article_id}")
async def delete_article(article_id: int, admin: str = Depends(verify_admin)):
    """Delete an article (Admin only)"""
    result = await db.articles.delete_one({"id": article_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    return {"message": "Article deleted successfully", "id": article_id}


# =============================================
# DYNAMIC SITEMAP
# =============================================

SITE_URL = "https://invest.euroadria.me"

# User-Agent parsing helper
def parse_device_type(user_agent: str) -> str:
    ua = user_agent.lower()
    if any(x in ua for x in ['mobile', 'android', 'iphone', 'ipod']):
        return 'mobile'
    if any(x in ua for x in ['ipad', 'tablet']):
        return 'tablet'
    return 'desktop'

STATIC_PAGES = [
    {"loc": "/", "priority": "1.0", "changefreq": "weekly"},
    {"loc": "/blog", "priority": "0.9", "changefreq": "daily"},
    {"loc": "/serbia-executive", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/serbia-executive/crypto-banking", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/serbia-executive/crypto-compliance", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/infrastruktur-radar", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/investment", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/investment/rechner", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/investment/simulation", "priority": "0.9", "changefreq": "weekly"},
    {"loc": "/investment/vergleich", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/contact", "priority": "0.8", "changefreq": "monthly"},
    {"loc": "/team", "priority": "0.7", "changefreq": "monthly"},
    {"loc": "/leistungen", "priority": "0.9", "changefreq": "monthly"},
    {"loc": "/events", "priority": "0.7", "changefreq": "weekly"},
    {"loc": "/immobilien/budva", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/niksic", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/podgorica", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/skadar-lake", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/immobilien/zabljak", "priority": "0.8", "changefreq": "weekly"},
    {"loc": "/impressum", "priority": "0.3", "changefreq": "yearly"},
    {"loc": "/datenschutz", "priority": "0.3", "changefreq": "yearly"},
    {"loc": "/agb", "priority": "0.3", "changefreq": "yearly"},
]

@api_router.get("/sitemap.xml")
async def generate_sitemap():
    """Generate dynamic sitemap from database articles + static pages"""
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    
    xml_parts = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml_parts.append('<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">')
    
    # Static pages
    for page in STATIC_PAGES:
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}{page["loc"]}</loc>')
        xml_parts.append(f'    <lastmod>{today}</lastmod>')
        xml_parts.append(f'    <changefreq>{page["changefreq"]}</changefreq>')
        xml_parts.append(f'    <priority>{page["priority"]}</priority>')
        xml_parts.append('  </url>')
    
    # Dynamic articles from DB
    articles = await db.articles.find({}, {"_id": 0, "slug": 1, "date": 1}).to_list(1000)
    for article in articles:
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}/blog/{article["slug"]}</loc>')
        xml_parts.append(f'    <lastmod>{article.get("date", today)}</lastmod>')
        xml_parts.append('    <changefreq>monthly</changefreq>')
        xml_parts.append('    <priority>0.7</priority>')
        xml_parts.append('  </url>')
    
    # Dynamic location profiles from DB
    locations = await db.locations.find({}, {"_id": 0, "city": 1}).to_list(100)
    for loc in locations:
        city_slug = loc["city"].replace(" ", "%20")
        xml_parts.append('  <url>')
        xml_parts.append(f'    <loc>{SITE_URL}/investment/standort/{city_slug}</loc>')
        xml_parts.append(f'    <lastmod>{today}</lastmod>')
        xml_parts.append('    <changefreq>weekly</changefreq>')
        xml_parts.append('    <priority>0.7</priority>')
        xml_parts.append('  </url>')
    
    xml_parts.append('</urlset>')
    
    return Response(
        content="\n".join(xml_parts), 
        media_type="application/xml",
        headers={"Content-Type": "application/xml; charset=utf-8"}
    )


# =============================================
# SITE SETTINGS ENDPOINTS
# =============================================

SETTINGS_FIXED_KEYS = [
    "praxisleitfaden_url",
    "budva_expose_url",
    "niksic_expose_url",
    "podgorica_expose_url",
    "skadar_lake_expose_url",
    "zabljak_expose_url"
]

@api_router.get("/settings/downloads")
async def get_download_settings():
    """Get all download URL settings (public)"""
    settings = await db.site_settings.find_one({"key": "downloads"}, {"_id": 0})
    if not settings:
        result = {k: "" for k in SETTINGS_FIXED_KEYS}
        result["custom_exposes"] = []
        return result
    result = {k: settings.get(k, "") for k in SETTINGS_FIXED_KEYS}
    result["custom_exposes"] = settings.get("custom_exposes", [])
    return result

@api_router.put("/admin/settings/downloads")
async def update_download_settings(settings: dict, admin: str = Depends(verify_admin)):
    """Update download URL settings (Admin only)"""
    update_data = {k: settings.get(k, "") for k in SETTINGS_FIXED_KEYS}
    update_data["custom_exposes"] = settings.get("custom_exposes", [])
    update_data["key"] = "downloads"
    await db.site_settings.update_one(
        {"key": "downloads"},
        {"$set": update_data},
        upsert=True
    )
    result = {k: update_data[k] for k in SETTINGS_FIXED_KEYS}
    result["custom_exposes"] = update_data["custom_exposes"]
    return result


# =============================================
# HOMEPAGE CONTENT SETTINGS
# =============================================

HOMEPAGE_DEFAULTS = {
    "hero_title": "Firmengründung, Aufenthalt & Investments in Montenegro und Serbien",
    "hero_subtitle": "EuroAdria ist Ihre Brücke zu erfolgreichen Investitionen, rechtssicherer Auswanderung und internationaler Unternehmensstrukturierung, sowohl in der Adria-Region als auch in Asien. Wir sind Ihr Trusted Advisor für alle unternehmerischen und privaten Vorhaben im Ausland.",
    "hero_cta_text": "Jetzt Beratung anfragen",
    "testimonial_image": "",
    "testimonial_image_position": 50,
    "testimonial_quote": "Dank EuroAdria konnte ich meine Firmengründung in Montenegro schnell, sicher und komplett stressfrei umsetzen. Ich habe mich bestens betreut gefühlt und kann EuroAdria jedem Unternehmer und Investor wärmstens empfehlen.",
    "testimonial_author": "Maximilian R., Unternehmer aus Deutschland",
    "stats_image": "https://images.unsplash.com/photo-1517048676732-d65bc937f952",
    "stats_image_position": 50,
    "cta_title": "Bereit für Ihre Investition?",
    "cta_subtitle": "Vereinbaren Sie ein unverbindliches Erstgespräch mit unseren Experten und entdecken Sie die Möglichkeiten am Balkan.",
    "trust_items": [
        {"title": "Vertrauenswürdig", "desc": "Referenziert in n-tv & RTL"},
        {"title": "Rendite-Fokus", "desc": "Zweistellige Zielrenditen"},
        {"title": "Expertise", "desc": "15+ Jahre Erfahrung"},
        {"title": "Sicherheit", "desc": "Asset Protection"}
    ]
}

@api_router.get("/settings/homepage")
async def get_homepage_settings():
    """Get homepage content settings (public)"""
    settings = await db.site_settings.find_one({"key": "homepage"}, {"_id": 0})
    if not settings:
        return HOMEPAGE_DEFAULTS
    result = {}
    for k, v in HOMEPAGE_DEFAULTS.items():
        saved = settings.get(k)
        # Use default if not saved or empty string
        if saved is None or saved == "":
            result[k] = v
        else:
            result[k] = saved
    return result

@api_router.put("/admin/settings/homepage")
async def update_homepage_settings(settings: dict, admin: str = Depends(verify_admin)):
    """Update homepage content settings (Admin only)"""
    update_data = {}
    for k in HOMEPAGE_DEFAULTS:
        if k in settings:
            update_data[k] = settings[k]
        else:
            update_data[k] = HOMEPAGE_DEFAULTS[k]
    update_data["key"] = "homepage"
    await db.site_settings.update_one(
        {"key": "homepage"},
        {"$set": update_data},
        upsert=True
    )
    return {k: update_data[k] for k in HOMEPAGE_DEFAULTS}


# =============================================
# LEGAL PAGES (Impressum & Datenschutz)
# =============================================

@api_router.get("/settings/legal/{page_type}")
async def get_legal_page(page_type: str):
    """Get legal page content (public). page_type: impressum, datenschutz or agb"""
    if page_type not in ("impressum", "datenschutz", "agb"):
        raise HTTPException(status_code=404, detail="Page not found")
    doc = await db.site_settings.find_one({"key": f"legal_{page_type}"}, {"_id": 0})
    if not doc or not doc.get("content"):
        return {"content": ""}
    return {"content": doc["content"]}

@api_router.put("/admin/settings/legal/{page_type}")
async def update_legal_page(page_type: str, data: dict, admin: str = Depends(verify_admin)):
    """Update legal page content (Admin only)"""
    if page_type not in ("impressum", "datenschutz", "agb"):
        raise HTTPException(status_code=404, detail="Page not found")
    content = data.get("content", "")
    await db.site_settings.update_one(
        {"key": f"legal_{page_type}"},
        {"$set": {"key": f"legal_{page_type}", "content": content}},
        upsert=True
    )
    return {"content": content}



# =============================================
# IMAGE UPLOAD ENDPOINT (Admin only)
# =============================================

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_IMAGE_DIMENSION = 1920  # Max width/height for optimization

def optimize_image(image_data: bytes, max_dimension: int = MAX_IMAGE_DIMENSION, quality: int = 85) -> bytes:
    """Optimize image: resize if too large, convert to WebP for better compression"""
    from PIL import Image
    img = Image.open(io.BytesIO(image_data))
    
    # Convert to RGB if necessary (for PNG with transparency)
    if img.mode in ('RGBA', 'P'):
        # Create white background
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'RGBA':
            background.paste(img, mask=img.split()[3])
        else:
            background.paste(img)
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Resize if too large
    width, height = img.size
    if width > max_dimension or height > max_dimension:
        ratio = min(max_dimension / width, max_dimension / height)
        new_size = (int(width * ratio), int(height * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
    
    # Save as WebP for better compression
    output = io.BytesIO()
    img.save(output, format='WEBP', quality=quality, optimize=True)
    return output.getvalue()


@api_router.post("/admin/upload")
async def upload_image(
    file: UploadFile = File(...),
    admin: str = Depends(verify_admin)
):
    """Upload and optimize an image (Admin only)"""
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Dateityp nicht erlaubt. Erlaubt: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Read file
    content = await file.read()
    
    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400, 
            detail=f"Datei zu groß. Maximum: {MAX_FILE_SIZE // (1024*1024)}MB"
        )
    
    try:
        # Optimize image
        optimized = optimize_image(content)
        
        # Generate unique filename
        unique_id = str(uuid.uuid4())[:8]
        original_name = Path(file.filename).stem
        # Clean filename
        clean_name = "".join(c for c in original_name if c.isalnum() or c in "-_")[:30]
        filename = f"{clean_name}_{unique_id}.webp"
        
        # Save file
        file_path = UPLOAD_DIR / filename
        with open(file_path, "wb") as f:
            f.write(optimized)
        
        # Return URL path
        url = f"/uploads/{filename}"
        
        # Calculate size reduction
        original_size = len(content)
        optimized_size = len(optimized)
        reduction = ((original_size - optimized_size) / original_size) * 100 if original_size > 0 else 0
        
        return {
            "url": url,
            "filename": filename,
            "originalSize": original_size,
            "optimizedSize": optimized_size,
            "reduction": f"{reduction:.1f}%"
        }
        
    except Exception as e:
        logger.error(f"Image upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Bildverarbeitung fehlgeschlagen: {str(e)}")


@api_router.get("/admin/uploads")
async def list_uploads(admin: str = Depends(verify_admin)):
    """List all uploaded images (Admin only)"""
    uploads = []
    for file_path in UPLOAD_DIR.glob("*"):
        if file_path.suffix.lower() in ALLOWED_EXTENSIONS or file_path.suffix.lower() == ".webp":
            stat = file_path.stat()
            uploads.append({
                "filename": file_path.name,
                "url": f"/uploads/{file_path.name}",
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat()
            })
    
    # Sort by creation date, newest first
    uploads.sort(key=lambda x: x["created"], reverse=True)
    return uploads


@api_router.delete("/admin/uploads/{filename}")
async def delete_upload(filename: str, admin: str = Depends(verify_admin)):
    """Delete an uploaded image (Admin only)"""
    file_path = UPLOAD_DIR / filename
    
    # Security check - prevent directory traversal
    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Ungültiger Dateiname")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Datei nicht gefunden")
    
    file_path.unlink()
    return {"message": "Datei gelöscht", "filename": filename}


@api_router.get("/admin/verify")
async def verify_admin_access(admin: str = Depends(verify_admin)):
    """Verify admin credentials"""
    return {"authenticated": True, "username": admin}


@api_router.post("/admin/seed-articles")
async def seed_articles(admin: str = Depends(verify_admin)):
    """Seed the database with initial articles (Admin only)"""
    # Check if articles already exist
    count = await db.articles.count_documents({})
    if count > 0:
        return {"message": f"Database already contains {count} articles. Skipping seed.", "seeded": False}
    
    # The pillar articles data will be sent from frontend during initial setup
    return {"message": "Use POST /api/admin/seed-articles-data with articles array to seed", "seeded": False}


@api_router.post("/admin/seed-articles-data")
async def seed_articles_data(articles: List[ArticleCreate], admin: str = Depends(verify_admin)):
    """Seed the database with provided articles data (Admin only)"""
    # Clear existing articles
    await db.articles.delete_many({})
    
    # Insert all articles with IDs
    for i, article in enumerate(articles):
        article_dict = article.model_dump()
        article_dict["id"] = 101 + i
        await db.articles.insert_one(article_dict)
    
    return {"message": f"Successfully seeded {len(articles)} articles", "seeded": True, "count": len(articles)}


@api_router.post("/admin/articles/bulk-import")
async def bulk_import_articles(file: UploadFile = File(...), admin: str = Depends(verify_admin)):
    """Bulk import articles from CSV, XLSX, or DOCX files"""
    import csv as csv_module
    
    filename = file.filename.lower()
    content = await file.read()
    
    articles_data = []
    
    try:
        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv_module.DictReader(io.StringIO(text), delimiter=';')
            if not reader.fieldnames or 'Titel' not in [f.strip() for f in reader.fieldnames]:
                reader = csv_module.DictReader(io.StringIO(text), delimiter=',')
            for row in reader:
                row = {k.strip(): v.strip() if v else '' for k, v in row.items() if k}
                if row.get('Titel'):
                    articles_data.append({
                        'title': row.get('Titel', ''),
                        'category': row.get('Kategorie', 'Allgemein'),
                        'excerpt': row.get('Excerpt', row.get('Beschreibung', '')),
                        'content': row.get('Content', row.get('Inhalt', '')),
                        'image': row.get('Bild', row.get('Bild-URL', row.get('Image', ''))),
                        'author': row.get('Autor', 'EuroAdria Corporate Solutions'),
                        'readTime': row.get('Lesezeit', '5 min'),
                        'downloadUrl': row.get('Download-URL', row.get('DownloadUrl', None)) or None,
                    })
                    
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Excel-Datei ist leer")
            headers = [str(h).strip() if h else '' for h in rows[0]]
            col_map = {}
            for i, h in enumerate(headers):
                hl = h.lower()
                if 'titel' in hl or 'title' in hl: col_map['title'] = i
                elif 'kategorie' in hl or 'category' in hl or 'cluster' in hl: col_map['category'] = i
                elif 'excerpt' in hl or 'beschreibung' in hl: col_map['excerpt'] = i
                elif 'content' in hl or 'inhalt' in hl: col_map['content'] = i
                elif 'bild' in hl or 'image' in hl: col_map['image'] = i
                elif 'autor' in hl or 'author' in hl: col_map['author'] = i
                elif 'lesezeit' in hl or 'readtime' in hl: col_map['readTime'] = i
                elif 'download' in hl: col_map['downloadUrl'] = i
            
            if 'title' not in col_map:
                raise HTTPException(status_code=400, detail="Spalte 'Titel' nicht gefunden. Erwartete Spalten: Titel, Kategorie, Excerpt, Content, Bild")
            
            for row in rows[1:]:
                title = str(row[col_map['title']]).strip() if col_map.get('title') is not None and row[col_map['title']] else ''
                if title:
                    articles_data.append({
                        'title': title,
                        'category': str(row[col_map.get('category', 0)] or 'Allgemein').strip() if col_map.get('category') is not None else 'Allgemein',
                        'excerpt': str(row[col_map.get('excerpt', 0)] or '').strip() if col_map.get('excerpt') is not None else '',
                        'content': str(row[col_map.get('content', 0)] or '').strip() if col_map.get('content') is not None else '',
                        'image': str(row[col_map.get('image', 0)] or '').strip() if col_map.get('image') is not None else '',
                        'author': str(row[col_map.get('author', 0)] or 'EuroAdria Corporate Solutions').strip() if col_map.get('author') is not None else 'EuroAdria Corporate Solutions',
                        'readTime': str(row[col_map.get('readTime', 0)] or '5 min').strip() if col_map.get('readTime') is not None else '5 min',
                        'downloadUrl': str(row[col_map.get('downloadUrl', 0)] or '').strip() if col_map.get('downloadUrl') is not None else None,
                    })
            wb.close()
                    
        elif filename.endswith('.docx'):
            from docx import Document
            doc = Document(io.BytesIO(content))
            current_article = None
            current_content = []
            current_category = 'Allgemein'
            
            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    if current_content:
                        current_content.append('')
                    continue
                    
                if para.style.name.startswith('Heading 1') or (para.runs and para.runs[0].bold and len(text) < 150 and not text.endswith('.')):
                    if current_article:
                        content_html = ''
                        for line in current_content:
                            if line == '':
                                continue
                            elif line.startswith('##'):
                                content_html += f'<h3>{line[2:].strip()}</h3>'
                            else:
                                content_html += f'<p>{line}</p>'
                        current_article['content'] = content_html
                        current_article['excerpt'] = current_content[0][:200] if current_content else ''
                        articles_data.append(current_article)
                    
                    if text.startswith('[') and ']' in text:
                        current_category = text[1:text.index(']')].strip()
                        title_part = text[text.index(']')+1:].strip()
                        if title_part:
                            current_article = {'title': title_part, 'category': current_category, 'content': '', 'excerpt': '', 'image': '', 'author': 'EuroAdria Corporate Solutions', 'readTime': '5 min', 'downloadUrl': None}
                        else:
                            current_article = None
                    else:
                        current_article = {'title': text, 'category': current_category, 'content': '', 'excerpt': '', 'image': '', 'author': 'EuroAdria Corporate Solutions', 'readTime': '5 min', 'downloadUrl': None}
                    current_content = []
                    
                elif para.style.name.startswith('Heading 2'):
                    current_content.append(f'##{text}')
                else:
                    current_content.append(text)
            
            if current_article:
                content_html = ''
                for line in current_content:
                    if line == '':
                        continue
                    elif line.startswith('##'):
                        content_html += f'<h3>{line[2:].strip()}</h3>'
                    else:
                        content_html += f'<p>{line}</p>'
                current_article['content'] = content_html
                current_article['excerpt'] = current_content[0][:200] if current_content else ''
                articles_data.append(current_article)
        else:
            raise HTTPException(status_code=400, detail="Nicht unterstütztes Format. Bitte CSV, XLSX oder DOCX verwenden.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fehler beim Parsen der Datei: {str(e)}")
    
    if not articles_data:
        raise HTTPException(status_code=400, detail="Keine Artikel in der Datei gefunden.")
    
    max_id_doc = await db.articles.find_one(sort=[("id", -1)])
    next_id = (max_id_doc["id"] + 1) if max_id_doc else 101
    
    import re
    def make_slug(title):
        slug = title.lower().strip()
        slug = slug.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
        slug = re.sub(r'[^a-z0-9]+', '-', slug)
        slug = slug.strip('-')
        return slug[:80]
    
    # Pre-load all existing slugs and titles for duplicate detection
    existing_articles = await db.articles.find({}, {"slug": 1, "title": 1, "_id": 0}).to_list(10000)
    existing_slugs = {a["slug"] for a in existing_articles}
    existing_titles = {a["title"].strip().lower() for a in existing_articles if a.get("title")}
    
    imported = []
    skipped = []
    file_seen_slugs = set()  # Track duplicates within the same file
    
    for art in articles_data:
        title = art['title'].strip()
        if not title:
            skipped.append({"title": "(leer)", "reason": "Kein Titel"})
            continue
        
        slug = make_slug(title)
        
        # Check 1: Duplicate slug in database
        if slug in existing_slugs:
            skipped.append({"title": title, "reason": f"Slug '{slug}' existiert bereits in der Datenbank"})
            continue
        
        # Check 2: Duplicate title in database (case-insensitive)
        if title.strip().lower() in existing_titles:
            skipped.append({"title": title, "reason": "Titel existiert bereits in der Datenbank"})
            continue
        
        # Check 3: Duplicate within the same import file
        if slug in file_seen_slugs:
            skipped.append({"title": title, "reason": "Duplikat innerhalb der Import-Datei"})
            continue
        
        file_seen_slugs.add(slug)
        
        category = art.get('category', 'Allgemein')
        
        article_doc = {
            "id": next_id,
            "cluster": category,
            "title": title,
            "slug": slug,
            "excerpt": art.get('excerpt', '')[:300],
            "content": art.get('content', ''),
            "image": art.get('image', ''),
            "category": category,
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "readTime": art.get('readTime', '5 min'),
            "featured": False,
            "author": art.get('author', 'EuroAdria Corporate Solutions'),
            "relatedArticles": [],
            "dueDiligenceBox": None,
            "expertTip": None,
            "downloadUrl": art.get('downloadUrl') or None,
            "sortOrder": 0,
            "imagePosition": 50
        }
        
        await db.articles.insert_one(article_doc)
        existing_slugs.add(slug)
        existing_titles.add(title.strip().lower())
        imported.append({"id": next_id, "title": title, "category": category, "slug": slug})
        next_id += 1
    
    result = {
        "message": f"{len(imported)} Artikel importiert, {len(skipped)} übersprungen",
        "count": len(imported),
        "skipped_count": len(skipped),
        "articles": imported,
    }
    if skipped:
        result["skipped"] = skipped
    
    return result


# =============================================
# COMMENT ENDPOINTS (Public & Admin)
# =============================================

@api_router.get("/comments/article/{article_id}")
async def get_approved_comments(article_id: int):
    """Get all approved comments for an article (Public)"""
    comments = await db.comments.find(
        {"articleId": article_id, "status": "approved"},
        {"_id": 0, "email": 0}  # Don't expose email publicly
    ).sort("createdAt", -1).to_list(100)
    return comments


@api_router.get("/comments/slug/{article_slug}")
async def get_approved_comments_by_slug(article_slug: str):
    """Get all approved comments for an article by slug (Public)"""
    comments = await db.comments.find(
        {"articleSlug": article_slug, "status": "approved"},
        {"_id": 0, "email": 0}  # Don't expose email publicly
    ).sort("createdAt", -1).to_list(100)
    return comments


@api_router.post("/comments")
async def create_comment(comment: CommentCreate, background_tasks: BackgroundTasks):
    """Create a new comment (requires moderation)"""
    # Get article info for email notification
    article = await db.articles.find_one({"id": comment.articleId}, {"title": 1})
    article_title = article.get("title", "Unknown Article") if article else "Unknown Article"
    
    comment_id = str(uuid.uuid4())
    comment_dict = {
        "id": comment_id,
        **comment.model_dump(),
        "status": "pending",  # pending, approved, rejected
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    
    await db.comments.insert_one(comment_dict)
    
    # Send notification email in background
    background_tasks.add_task(send_notification_email, comment_dict, article_title)
    
    return {"message": "Comment submitted for moderation", "id": comment_id}


@api_router.get("/admin/comments")
async def get_all_comments(
    status: Optional[str] = None,
    admin: str = Depends(verify_admin)
):
    """Get all comments with optional status filter (Admin only)"""
    query = {}
    if status:
        query["status"] = status
    
    comments = await db.comments.find(query, {"_id": 0}).sort("createdAt", -1).to_list(500)
    
    # Enrich with article titles
    for comment in comments:
        article = await db.articles.find_one({"id": comment.get("articleId")}, {"title": 1})
        comment["articleTitle"] = article.get("title", "Unknown") if article else "Unknown"
    
    return comments


@api_router.get("/admin/comments/stats")
async def get_comments_stats(admin: str = Depends(verify_admin)):
    """Get comment statistics (Admin only)"""
    total = await db.comments.count_documents({})
    pending = await db.comments.count_documents({"status": "pending"})
    approved = await db.comments.count_documents({"status": "approved"})
    rejected = await db.comments.count_documents({"status": "rejected"})
    
    return {
        "total": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected
    }


@api_router.put("/admin/comments/{comment_id}/approve")
async def approve_comment(comment_id: str, admin: str = Depends(verify_admin)):
    """Approve a comment (Admin only)"""
    result = await db.comments.update_one(
        {"id": comment_id},
        {"$set": {"status": "approved"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Comment not found")
    return {"message": "Comment approved", "id": comment_id}


@api_router.put("/admin/comments/{comment_id}/reject")
async def reject_comment(comment_id: str, admin: str = Depends(verify_admin)):
    """Reject a comment (Admin only)"""
    result = await db.comments.update_one(
        {"id": comment_id},
        {"$set": {"status": "rejected"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Comment not found")
    return {"message": "Comment rejected", "id": comment_id}


@api_router.delete("/admin/comments/{comment_id}")
async def delete_comment(comment_id: str, admin: str = Depends(verify_admin)):
    """Delete a comment (Admin only)"""
    result = await db.comments.delete_one({"id": comment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Comment not found")
    return {"message": "Comment deleted", "id": comment_id}


# ============================================
# REGIONS API (Public + Admin)
# ============================================

@api_router.get("/regions")
async def get_all_regions():
    """Get all regions (Public)"""
    regions = await db.regions.find({}, {"_id": 0}).to_list(100)
    return regions


@api_router.get("/regions/{slug}")
async def get_region(slug: str):
    """Get a region by slug (Public)"""
    region = await db.regions.find_one({"slug": slug}, {"_id": 0})
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")
    return region


@api_router.get("/admin/regions")
async def get_admin_regions(admin: str = Depends(verify_admin)):
    """Get all regions with full details (Admin only)"""
    regions = await db.regions.find({}, {"_id": 0}).to_list(100)
    return regions


@api_router.post("/admin/regions")
async def create_region(region: RegionCreate, admin: str = Depends(verify_admin)):
    """Create a new region (Admin only)"""
    # Check if slug already exists
    existing = await db.regions.find_one({"slug": region.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Region with this slug already exists")
    
    region_dict = region.model_dump()
    region_dict["id"] = str(uuid.uuid4())
    region_dict["createdAt"] = datetime.now(timezone.utc).isoformat()
    region_dict["updatedAt"] = datetime.now(timezone.utc).isoformat()
    
    await db.regions.insert_one(region_dict)
    
    # Remove _id from response
    region_dict.pop("_id", None)
    return region_dict


@api_router.put("/admin/regions/{slug}")
async def update_region(slug: str, region_update: RegionUpdate, admin: str = Depends(verify_admin)):
    """Update a region (Admin only)"""
    update_data = {k: v for k, v in region_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data["updatedAt"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.regions.update_one(
        {"slug": slug},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Region not found")
    
    # Return updated region
    updated_region = await db.regions.find_one({"slug": slug}, {"_id": 0})
    return updated_region


@api_router.delete("/admin/regions/{slug}")
async def delete_region(slug: str, admin: str = Depends(verify_admin)):
    """Delete a region (Admin only)"""
    result = await db.regions.delete_one({"slug": slug})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Region not found")
    return {"message": "Region deleted", "slug": slug}


@api_router.post("/admin/regions/{slug}/apartments")
async def add_apartment_to_region(slug: str, apartment: RegionApartment, admin: str = Depends(verify_admin)):
    """Add an apartment to a region (Admin only)"""
    region = await db.regions.find_one({"slug": slug})
    if not region:
        raise HTTPException(status_code=404, detail="Region not found")
    
    apartment_dict = apartment.model_dump()
    apartment_dict["id"] = str(uuid.uuid4())
    
    await db.regions.update_one(
        {"slug": slug},
        {"$push": {"apartments": apartment_dict}}
    )
    
    return {"message": "Apartment added", "apartment": apartment_dict}


@api_router.delete("/admin/regions/{slug}/apartments/{apartment_id}")
async def remove_apartment_from_region(slug: str, apartment_id: str, admin: str = Depends(verify_admin)):
    """Remove an apartment from a region (Admin only)"""
    result = await db.regions.update_one(
        {"slug": slug},
        {"$pull": {"apartments": {"id": apartment_id}}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Region not found")
    
    return {"message": "Apartment removed", "apartment_id": apartment_id}


# =============================================
# PAGE CMS API (Public + Admin)
# =============================================

# Default page templates for initialization
DEFAULT_PAGES = {
    "home": {
        "title": "Home",
        "metaTitle": "EuroAdria - Investment & Business Beratung Adria",
        "metaDescription": "Professionelle Beratung für Investments, Business und Lifestyle an der Adria. Schwerpunkt Montenegro, Serbien und Balkan-Region.",
        "sections": [
            {
                "id": "hero",
                "type": "hero",
                "data": {
                    "title": "Investieren an der Adria",
                    "subtitle": "Ihr Partner für nachhaltige Investments in Montenegro & Serbien",
                    "ctaText": "Jetzt Beratung anfragen",
                    "ctaLink": "/contact"
                }
            },
            {
                "id": "why-balkan",
                "type": "cards",
                "title": "Warum Balkan statt EU?",
                "data": {
                    "cards": [
                        {"id": "1", "title": "Niedrige Steuern", "description": "Körperschaftssteuer nur 9% in Montenegro, 15% in Serbien", "icon": "percent"},
                        {"id": "2", "title": "EU-Beitritt", "description": "Montenegro und Serbien auf dem Weg in die EU - Wertsteigerungspotenzial", "icon": "flag"},
                        {"id": "3", "title": "Wachstumsmarkt", "description": "Tourismus wächst jährlich um 15-20%, Immobilienpreise steigen", "icon": "trending-up"},
                        {"id": "4", "title": "Strategische Lage", "description": "Tor zwischen Ost und West, neue Infrastrukturprojekte", "icon": "map-pin"}
                    ]
                }
            },
            {
                "id": "task-force",
                "type": "cards",
                "title": "Warum Task Force Adria?",
                "data": {
                    "cards": [
                        {"id": "1", "title": "Lokale Expertise", "description": "Unser Team lebt und arbeitet vor Ort", "icon": "users"},
                        {"id": "2", "title": "Rechtssicherheit", "description": "Due Diligence und rechtliche Begleitung inklusive", "icon": "shield"},
                        {"id": "3", "title": "Netzwerk", "description": "Direkter Zugang zu Off-Market Deals", "icon": "network"},
                        {"id": "4", "title": "Rundum-Service", "description": "Von der Suche bis zum Notar - alles aus einer Hand", "icon": "check-circle"}
                    ]
                }
            }
        ]
    },
    "team": {
        "title": "Über uns",
        "metaTitle": "Team - EuroAdria Investment Beratung",
        "metaDescription": "Lernen Sie unser erfahrenes Team kennen. Holger Kuhlmann und Milena Bubanja - Ihre Experten für Investments an der Adria.",
        "sections": [
            {
                "id": "hero",
                "type": "hero",
                "data": {
                    "title": "Unser Team",
                    "subtitle": "Erfahrung trifft lokale Expertise"
                }
            },
            {
                "id": "team-members",
                "type": "team",
                "data": {
                    "members": [
                        {
                            "id": "holger",
                            "name": "Holger Kuhlmann",
                            "title": "Berater & Leitung DACH",
                            "description": "Ich glaube daran, dass nachhaltige Projekte und solide Strukturen die beste Basis für langfristigen Erfolg sind.",
                            "image": "/holger-kuhlmann.jpg",
                            "email": "holger@euroadria.me"
                        },
                        {
                            "id": "milena",
                            "name": "Milena Bubanja",
                            "title": "Co-Founderin und Geschäftsführerin",
                            "subtitle": "Public Affairs und Balkan Relations",
                            "description": "Nachhaltige Ergebnisse entstehen dort, wo lokale Realität und europäische Standards sauber zusammengeführt werden.",
                            "image": "/milena-bubanja.jpg",
                            "email": "milena@euroadria.me"
                        }
                    ]
                }
            },
            {
                "id": "about-text",
                "type": "text",
                "title": "Unsere Mission",
                "content": "<p>EuroAdria wurde gegründet, um deutschsprachigen Investoren einen sicheren und professionellen Zugang zu den wachsenden Märkten Montenegros und Serbiens zu bieten.</p><p>Wir kombinieren deutsche Gründlichkeit mit lokalem Know-how und begleiten Sie von der ersten Idee bis zum erfolgreichen Abschluss.</p>"
            }
        ]
    },
    "contact": {
        "title": "Kontakt",
        "metaTitle": "Kontakt - EuroAdria Investment Beratung",
        "metaDescription": "Kontaktieren Sie uns für eine unverbindliche Erstberatung zu Investments in Montenegro und Serbien.",
        "sections": [
            {
                "id": "hero",
                "type": "hero",
                "data": {
                    "title": "Kontakt",
                    "subtitle": "Wir freuen uns auf Ihre Anfrage"
                }
            },
            {
                "id": "contact-info",
                "type": "contact",
                "data": {
                    "email": "info@euroadria.me",
                    "phone": "+382 69 123 456",
                    "address": "Podgorica, Montenegro",
                    "hours": "Mo-Fr: 9:00 - 18:00 Uhr"
                }
            }
        ]
    }
}


@api_router.get("/pages")
async def get_all_pages():
    """Get all pages (Public)"""
    pages = await db.pages.find({}, {"_id": 0}).to_list(100)
    return pages


@api_router.get("/pages/{slug}")
async def get_page(slug: str):
    """Get a page by slug (Public) - returns default if not in DB"""
    page = await db.pages.find_one({"slug": slug}, {"_id": 0})
    
    if not page:
        # Return default template if exists
        if slug in DEFAULT_PAGES:
            return {
                "slug": slug,
                "id": f"default-{slug}",
                **DEFAULT_PAGES[slug],
                "isDefault": True
            }
        raise HTTPException(status_code=404, detail="Page not found")
    
    return page


@api_router.get("/admin/pages")
async def get_admin_pages(admin: str = Depends(verify_admin)):
    """Get all pages with defaults (Admin only)"""
    pages = await db.pages.find({}, {"_id": 0}).to_list(100)
    
    # Add default pages that aren't in DB yet
    existing_slugs = {p["slug"] for p in pages}
    for slug, default_data in DEFAULT_PAGES.items():
        if slug not in existing_slugs:
            pages.append({
                "slug": slug,
                "id": f"default-{slug}",
                **default_data,
                "isDefault": True
            })
    
    return pages


@api_router.post("/admin/pages")
async def create_page(page: PageCreate, admin: str = Depends(verify_admin)):
    """Create a new page (Admin only)"""
    existing = await db.pages.find_one({"slug": page.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Page with this slug already exists")
    
    page_dict = page.model_dump()
    page_dict["id"] = str(uuid.uuid4())
    page_dict["updatedAt"] = datetime.now(timezone.utc).isoformat()
    
    await db.pages.insert_one(page_dict)
    page_dict.pop("_id", None)
    
    return page_dict


@api_router.put("/admin/pages/{slug}")
async def update_page(slug: str, page_update: PageUpdate, admin: str = Depends(verify_admin)):
    """Update or create a page (Admin only) - upsert from default"""
    update_data = {k: v for k, v in page_update.model_dump().items() if v is not None}
    update_data["updatedAt"] = datetime.now(timezone.utc).isoformat()
    
    existing = await db.pages.find_one({"slug": slug})
    
    if not existing:
        # Create from default template
        if slug in DEFAULT_PAGES:
            new_page = {
                "slug": slug,
                "id": str(uuid.uuid4()),
                **DEFAULT_PAGES[slug],
                **update_data
            }
            await db.pages.insert_one(new_page)
            new_page.pop("_id", None)
            return new_page
        else:
            raise HTTPException(status_code=404, detail="Page not found")
    
    await db.pages.update_one({"slug": slug}, {"$set": update_data})
    
    updated = await db.pages.find_one({"slug": slug}, {"_id": 0})
    return updated


@api_router.delete("/admin/pages/{slug}")
async def delete_page(slug: str, admin: str = Depends(verify_admin)):
    """Delete a page (resets to default if available) (Admin only)"""
    result = await db.pages.delete_one({"slug": slug})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Page not found or is default")
    
    return {"message": "Page deleted (reset to default)", "slug": slug}


# =============================================
# INVESTMENT INTELLIGENCE API
# =============================================

from investment_models import (
    LocationBase, LocationCreate, LocationUpdate, Location,
    InfrastructureProjectBase, InfrastructureProjectCreate, InfrastructureProjectUpdate, InfrastructureProject,
    OpportunityZoneBase, OpportunityZoneCreate, OpportunityZoneUpdate, OpportunityZone,
    ROICalculation, ROIResult,
    SimulationInput, SimulationResult, calculate_simulation,
    MarketCheckInput, MarketCheckResult, MarketCheckWarning,
    PredictiveScoreResult,
    INFRA_TYPE_WEIGHTS, INFRA_STATUS_MULTIPLIER,
    calculate_investment_score, calculate_roi,
    SEED_LOCATIONS, SEED_INFRASTRUCTURE, SEED_ZONES
)
import math


def calculate_distance_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two points using Haversine formula"""
    R = 6371  # Earth's radius in km
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c


async def calculate_infrastructure_boost(lat: float, lng: float) -> float:
    """
    Predictive Infrastructure Score Boost.
    Gewichtung: Flughafen +15%, Autobahn +10%, Klinik +8%
    Statusfaktor: Fertig 100%, Im Bau 60%, Geplant 20%
    """
    projects = await db.infrastructure_projects.find({}, {"_id": 0}).to_list(1000)
    
    boost = 0
    for project in projects:
        distance = calculate_distance_km(lat, lng, project["latitude"], project["longitude"])
        if distance <= project["impact_radius_km"]:
            # Entfernungsfaktor: Linear abnehmend (1.0 = direkt, 0.0 = am Rand)
            impact_factor = 1 - (distance / project["impact_radius_km"])
            
            # Typ-basierte Gewichtung (aus investment_models.py)
            type_boost = INFRA_TYPE_WEIGHTS.get(project["type"], 5.0)
            
            # Status-Multiplikator
            status_mult = INFRA_STATUS_MULTIPLIER.get(project["status"], 0.5)
            
            boost += type_boost * impact_factor * status_mult
    
    return min(boost, 25)  # Cap bei 25 Punkten


# ---- LOCATIONS API ----

@api_router.get("/locations")
async def get_all_locations():
    """Get all investment locations with calculated scores"""
    locations = await db.locations.find({}, {"_id": 0}).to_list(1000)
    
    # Recalculate scores for each location
    for loc in locations:
        infra_boost = await calculate_infrastructure_boost(loc["latitude"], loc["longitude"])
        adjusted_infra = min(loc["infrastructure_score"] + infra_boost, 100)
        
        loc["investment_score"] = calculate_investment_score(
            adjusted_infra,
            loc["tourism_growth"],
            loc["rental_yield"],
            loc["price_growth"]
        )
        loc["infrastructure_boost"] = round(infra_boost, 1)
    
    # Sort by investment score
    locations.sort(key=lambda x: x["investment_score"], reverse=True)
    
    return locations


@api_router.get("/locations/{city}")
async def get_location(city: str):
    """Get a specific location by city name"""
    location = await db.locations.find_one(
        {"city": {"$regex": f"^{city}$", "$options": "i"}},
        {"_id": 0}
    )
    
    if not location:
        raise HTTPException(status_code=404, detail="Location not found")
    
    # Calculate infrastructure boost
    infra_boost = await calculate_infrastructure_boost(location["latitude"], location["longitude"])
    adjusted_infra = min(location["infrastructure_score"] + infra_boost, 100)
    
    location["investment_score"] = calculate_investment_score(
        adjusted_infra,
        location["tourism_growth"],
        location["rental_yield"],
        location["price_growth"]
    )
    location["infrastructure_boost"] = round(infra_boost, 1)
    
    # Get related blog articles
    related_articles = await db.articles.find(
        {"$or": [
            {"title": {"$regex": city, "$options": "i"}},
            {"content": {"$regex": city, "$options": "i"}},
            {"cluster": {"$regex": location["country"], "$options": "i"}}
        ]},
        {"_id": 0, "id": 1, "title": 1, "slug": 1, "excerpt": 1, "image": 1}
    ).limit(5).to_list(5)
    
    location["related_articles"] = related_articles
    
    return location


@api_router.get("/locations/compare/{cities}")
async def compare_locations(cities: str):
    """Compare multiple locations (comma-separated city names)"""
    city_list = [c.strip() for c in cities.split(",")]
    
    comparisons = []
    for city in city_list:
        location = await db.locations.find_one(
            {"city": {"$regex": f"^{city}$", "$options": "i"}},
            {"_id": 0}
        )
        if location:
            infra_boost = await calculate_infrastructure_boost(location["latitude"], location["longitude"])
            adjusted_infra = min(location["infrastructure_score"] + infra_boost, 100)
            
            location["investment_score"] = calculate_investment_score(
                adjusted_infra,
                location["tourism_growth"],
                location["rental_yield"],
                location["price_growth"]
            )
            comparisons.append(location)
    
    return comparisons


@api_router.post("/admin/locations")
async def create_location(location: LocationCreate, admin: str = Depends(verify_admin)):
    """Create a new investment location"""
    existing = await db.locations.find_one({"city": location.city})
    if existing:
        raise HTTPException(status_code=400, detail="Location already exists")
    
    loc_dict = location.model_dump()
    loc_dict["id"] = str(uuid.uuid4())
    loc_dict["investment_score"] = calculate_investment_score(
        location.infrastructure_score,
        location.tourism_growth,
        location.rental_yield,
        location.price_growth
    )
    loc_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.locations.insert_one(loc_dict)
    loc_dict.pop("_id", None)
    
    return loc_dict


@api_router.put("/admin/locations/{city}")
async def update_location(city: str, update: LocationUpdate, admin: str = Depends(verify_admin)):
    """Update an investment location"""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.locations.update_one(
        {"city": {"$regex": f"^{city}$", "$options": "i"}},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    
    updated = await db.locations.find_one(
        {"city": {"$regex": f"^{city}$", "$options": "i"}},
        {"_id": 0}
    )
    
    # Recalculate score
    updated["investment_score"] = calculate_investment_score(
        updated["infrastructure_score"],
        updated["tourism_growth"],
        updated["rental_yield"],
        updated["price_growth"]
    )
    
    return updated


@api_router.delete("/admin/locations/{city}")
async def delete_location(city: str, admin: str = Depends(verify_admin)):
    """Delete an investment location"""
    result = await db.locations.delete_one({"city": {"$regex": f"^{city}$", "$options": "i"}})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"message": "Location deleted", "city": city}


# ---- INFRASTRUCTURE API ----

@api_router.get("/infrastructure")
async def get_all_infrastructure():
    """Get all infrastructure projects"""
    projects = await db.infrastructure_projects.find({}, {"_id": 0}).to_list(1000)
    return projects


@api_router.get("/infrastructure/{project_id}")
async def get_infrastructure_project(project_id: str):
    """Get a specific infrastructure project"""
    project = await db.infrastructure_projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@api_router.post("/admin/infrastructure")
async def create_infrastructure(project: InfrastructureProjectCreate, admin: str = Depends(verify_admin)):
    """Create a new infrastructure project"""
    proj_dict = project.model_dump()
    proj_dict["id"] = str(uuid.uuid4())
    proj_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.infrastructure_projects.insert_one(proj_dict)
    proj_dict.pop("_id", None)
    
    return proj_dict


@api_router.put("/admin/infrastructure/{project_id}")
async def update_infrastructure(project_id: str, update: InfrastructureProjectUpdate, admin: str = Depends(verify_admin)):
    """Update an infrastructure project"""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.infrastructure_projects.update_one({"id": project_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    updated = await db.infrastructure_projects.find_one({"id": project_id}, {"_id": 0})
    return updated


@api_router.delete("/admin/infrastructure/{project_id}")
async def delete_infrastructure(project_id: str, admin: str = Depends(verify_admin)):
    """Delete an infrastructure project"""
    result = await db.infrastructure_projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Project deleted", "id": project_id}


# ---- OPPORTUNITY ZONES API ----

@api_router.get("/zones")
async def get_all_zones():
    """Get all opportunity zones"""
    zones = await db.opportunity_zones.find({}, {"_id": 0}).to_list(1000)
    return zones


@api_router.get("/zones/{zone_id}")
async def get_zone(zone_id: str):
    """Get a specific opportunity zone"""
    zone = await db.opportunity_zones.find_one({"id": zone_id}, {"_id": 0})
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    return zone


@api_router.post("/admin/zones")
async def create_zone(zone: OpportunityZoneCreate, admin: str = Depends(verify_admin)):
    """Create a new opportunity zone"""
    zone_dict = zone.model_dump()
    zone_dict["id"] = str(uuid.uuid4())
    
    await db.opportunity_zones.insert_one(zone_dict)
    zone_dict.pop("_id", None)
    
    return zone_dict


@api_router.put("/admin/zones/{zone_id}")
async def update_zone(zone_id: str, update: OpportunityZoneUpdate, admin: str = Depends(verify_admin)):
    """Update an opportunity zone"""
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.opportunity_zones.update_one({"id": zone_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Zone not found")
    
    updated = await db.opportunity_zones.find_one({"id": zone_id}, {"_id": 0})
    return updated


@api_router.delete("/admin/zones/{zone_id}")
async def delete_zone(zone_id: str, admin: str = Depends(verify_admin)):
    """Delete an opportunity zone"""
    result = await db.opportunity_zones.delete_one({"id": zone_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Zone not found")
    return {"message": "Zone deleted", "id": zone_id}


# ---- ROI CALCULATOR API ----

@api_router.post("/calculator/roi", response_model=ROIResult)
async def calculate_property_roi(calc: ROICalculation):
    """Calculate ROI for a property investment"""
    return calculate_roi(calc)


# ---- DYNAMIC SIMULATION (IRR & CASHFLOW-PROGNOSE) ----

@api_router.post("/calculator/simulation", response_model=SimulationResult)
async def run_investment_simulation(inp: SimulationInput):
    """
    10-Jahres Investment-Simulation mit IRR, NPV und kumuliertem Cashflow.
    Berechnet Eigenkapital-ROI, Wertsteigerung und Break-Even.
    """
    return calculate_simulation(inp)


# ---- PREDICTIVE INFRASTRUCTURE SCORE ----

@api_router.get("/locations/{city}/predictive-score")
async def get_predictive_score(city: str):
    """
    Prädiktiver Investment-Score für einen Standort.
    Zeigt aktuellen Score, Infrastructure-Boost und Was-wäre-wenn Szenarien.
    """
    location = await db.locations.find_one({"city": city}, {"_id": 0})
    if not location:
        raise HTTPException(status_code=404, detail=f"Standort '{city}' nicht gefunden")
    
    projects = await db.infrastructure_projects.find({}, {"_id": 0}).to_list(1000)
    
    # Aktuelle Infrastruktur-Boost Berechnung
    current_boost = await calculate_infrastructure_boost(location["latitude"], location["longitude"])
    adjusted_infra = min(location["infrastructure_score"] + current_boost, 100)
    current_score = calculate_investment_score(
        adjusted_infra, location["tourism_growth"],
        location["rental_yield"], location["price_growth"]
    )
    
    # Nahe Projekte finden
    nearby = []
    for proj in projects:
        dist = calculate_distance_km(
            location["latitude"], location["longitude"],
            proj["latitude"], proj["longitude"]
        )
        if dist <= proj["impact_radius_km"] * 1.5:  # 1.5x Radius für "in der Nähe"
            nearby.append({
                "project_name": proj["project_name"],
                "type": proj["type"],
                "status": proj["status"],
                "distance_km": round(dist, 1),
                "impact_radius_km": proj["impact_radius_km"],
                "within_range": dist <= proj["impact_radius_km"],
                "type_weight": INFRA_TYPE_WEIGHTS.get(proj["type"], 5.0),
                "status_multiplier": INFRA_STATUS_MULTIPLIER.get(proj["status"], 0.5)
            })
    
    # Was-wäre-wenn Szenarien: Jedes geplante/im-Bau Projekt auf "fertig" setzen
    scenarios = []
    for proj in projects:
        if proj["status"] in ("planned", "construction"):
            dist = calculate_distance_km(
                location["latitude"], location["longitude"],
                proj["latitude"], proj["longitude"]
            )
            if dist <= proj["impact_radius_km"]:
                # Simuliere: Projekt wird fertiggestellt
                impact_factor = 1 - (dist / proj["impact_radius_km"])
                type_w = INFRA_TYPE_WEIGHTS.get(proj["type"], 5.0)
                
                current_contribution = type_w * impact_factor * INFRA_STATUS_MULTIPLIER.get(proj["status"], 0.5)
                completed_contribution = type_w * impact_factor * 1.0  # Status "built"
                delta = completed_contribution - current_contribution
                
                # Neuer Score mit fertiggestelltem Projekt
                new_boost = current_boost + delta
                new_infra = min(location["infrastructure_score"] + new_boost, 100)
                new_score = calculate_investment_score(
                    new_infra, location["tourism_growth"],
                    location["rental_yield"], location["price_growth"]
                )
                
                scenarios.append({
                    "project_name": proj["project_name"],
                    "current_status": proj["status"],
                    "simulated_status": "built",
                    "score_impact": round(new_score - current_score, 1),
                    "predicted_score": round(new_score, 1),
                    "distance_km": round(dist, 1)
                })
    
    # Gesamtszenario: Alle geplanten Projekte fertig
    total_boost_all = 0
    for proj in projects:
        dist = calculate_distance_km(
            location["latitude"], location["longitude"],
            proj["latitude"], proj["longitude"]
        )
        if dist <= proj["impact_radius_km"]:
            impact_factor = 1 - (dist / proj["impact_radius_km"])
            type_w = INFRA_TYPE_WEIGHTS.get(proj["type"], 5.0)
            total_boost_all += type_w * impact_factor * 1.0  # Alles "built"
    
    max_infra = min(location["infrastructure_score"] + total_boost_all, 100)
    predicted_max = calculate_investment_score(
        max_infra, location["tourism_growth"],
        location["rental_yield"], location["price_growth"]
    )
    
    return PredictiveScoreResult(
        city=city,
        current_score=round(current_score, 1),
        predicted_score=round(predicted_max, 1),
        score_delta=round(predicted_max - current_score, 1),
        infrastructure_boost=round(current_boost, 1),
        nearby_projects=sorted(nearby, key=lambda x: x["distance_km"]),
        scenarios=sorted(scenarios, key=lambda x: x["score_impact"], reverse=True)
    )


# ---- MARKET-DATA-VALIDATION ----

@api_router.post("/v1/market-check", response_model=MarketCheckResult)
async def market_data_check(inp: MarketCheckInput):
    """
    Prüft Benutzereingaben gegen Marktdurchschnitte.
    Warning-Flag bei Abweichung > 15%.
    """
    warnings = []
    market_avgs = {}
    
    # Marktdaten aus der Datenbank laden
    location = None
    if inp.city:
        location = await db.locations.find_one({"city": inp.city}, {"_id": 0})
    
    if not location and inp.country:
        # Durchschnitt aller Standorte im Land
        locations = await db.locations.find({"country": inp.country}, {"_id": 0}).to_list(1000)
        if locations:
            location = {
                "city": f"Durchschnitt {inp.country}",
                "price_per_m2": sum(l["price_per_m2"] for l in locations) / len(locations),
                "rental_yield": sum(l["rental_yield"] for l in locations) / len(locations),
                "price_growth": sum(l["price_growth"] for l in locations) / len(locations),
            }
    
    if not location:
        return MarketCheckResult(
            city=inp.city,
            market_data_available=False,
            warnings=[],
            market_averages={},
            risk_level="unknown",
            summary="Keine Marktdaten für diesen Standort verfügbar."
        )
    
    market_avgs = {
        "price_per_m2": round(location.get("price_per_m2", 0), 2),
        "rental_yield": round(location.get("rental_yield", 0), 2),
        "price_growth": round(location.get("price_growth", 0), 2),
    }
    
    # Abweichungen berechnen (Monatsmiete → Mietrendite ableiten)
    if inp.price_per_m2 and location.get("price_per_m2"):
        avg = location["price_per_m2"]
        dev = ((inp.price_per_m2 - avg) / avg) * 100
        severity = "critical" if abs(dev) > 30 else "warning" if abs(dev) > 15 else "info"
        direction = "über" if dev > 0 else "unter"
        warnings.append(MarketCheckWarning(
            field="price_per_m2",
            user_value=inp.price_per_m2,
            market_average=round(avg, 2),
            deviation_percent=round(dev, 1),
            severity=severity,
            message=f"Ihr Preis von {inp.price_per_m2:.0f} EUR/m² liegt {abs(dev):.0f}% {direction} dem Marktdurchschnitt von {avg:.0f} EUR/m²."
        ))
    
    if inp.rental_yield and location.get("rental_yield"):
        avg = location["rental_yield"]
        dev = ((inp.rental_yield - avg) / avg) * 100
        severity = "critical" if abs(dev) > 30 else "warning" if abs(dev) > 15 else "info"
        direction = "über" if dev > 0 else "unter"
        warnings.append(MarketCheckWarning(
            field="rental_yield",
            user_value=inp.rental_yield,
            market_average=round(avg, 2),
            deviation_percent=round(dev, 1),
            severity=severity,
            message=f"Ihre Mietrendite von {inp.rental_yield:.1f}% liegt {abs(dev):.0f}% {direction} dem Marktdurchschnitt von {avg:.1f}%."
        ))
    
    # Mietpreis pro m² berechnen wenn Kaufpreis und Größe gegeben
    if inp.monthly_rent_per_m2 and inp.price_per_m2:
        implied_yield = (inp.monthly_rent_per_m2 * 12 / inp.price_per_m2) * 100
        avg_yield = location.get("rental_yield", 5.0)
        dev = ((implied_yield - avg_yield) / avg_yield) * 100
        severity = "critical" if abs(dev) > 30 else "warning" if abs(dev) > 15 else "info"
        market_avgs["implied_yield"] = round(implied_yield, 2)
        if abs(dev) > 15:
            warnings.append(MarketCheckWarning(
                field="implied_yield",
                user_value=round(implied_yield, 2),
                market_average=round(avg_yield, 2),
                deviation_percent=round(dev, 1),
                severity=severity,
                message=f"Ihre implizierte Rendite von {implied_yield:.1f}% weicht {abs(dev):.0f}% vom Marktdurchschnitt ab."
            ))
    
    # Risikobewertung
    critical_count = sum(1 for w in warnings if w.severity == "critical")
    warning_count = sum(1 for w in warnings if w.severity == "warning")
    
    if critical_count > 0:
        risk_level = "high"
        summary = f"Achtung: {critical_count} kritische Abweichung(en) vom Marktdurchschnitt. Bitte Eingaben prüfen."
    elif warning_count > 0:
        risk_level = "medium"
        summary = f"{warning_count} Eingabe(n) weichen über 15% vom Markt ab. Empfehlung: Marktforschung vertiefen."
    else:
        risk_level = "low"
        summary = "Ihre Eingaben liegen im Rahmen der aktuellen Marktdaten."
    
    return MarketCheckResult(
        city=inp.city or location.get("city"),
        market_data_available=True,
        warnings=warnings,
        market_averages=market_avgs,
        risk_level=risk_level,
        summary=summary
    )


# ---- PDF EXPOSÉ-GENERATOR ----

@api_router.post("/calculator/expose-pdf")
async def generate_expose_pdf(inp: SimulationInput):
    """
    Generiert ein professionelles PDF-Exposé mit ROI-Daten,
    10-Jahres-Prognose und Investment-Kennzahlen.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from starlette.responses import Response
    
    # Simulation berechnen
    result = calculate_simulation(inp)
    
    # PDF erstellen
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=20*mm, rightMargin=20*mm,
                           topMargin=20*mm, bottomMargin=20*mm)
    
    # Farben
    dark = HexColor('#04151F')
    gray = HexColor('#6B7280')
    light_gray = HexColor('#F3F4F6')
    
    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Title_EA', parent=styles['Title'],
                              fontSize=22, textColor=dark, spaceAfter=4*mm))
    styles.add(ParagraphStyle('Subtitle_EA', parent=styles['Normal'],
                              fontSize=11, textColor=gray, spaceAfter=8*mm))
    styles.add(ParagraphStyle('Section_EA', parent=styles['Heading2'],
                              fontSize=14, textColor=dark, spaceBefore=8*mm, spaceAfter=4*mm,
                              borderWidth=0, borderPadding=0))
    styles.add(ParagraphStyle('Body_EA', parent=styles['Normal'],
                              fontSize=10, textColor=dark, leading=14))
    styles.add(ParagraphStyle('Small_EA', parent=styles['Normal'],
                              fontSize=8, textColor=gray))
    styles.add(ParagraphStyle('KPI_Value', parent=styles['Normal'],
                              fontSize=16, textColor=dark, alignment=TA_CENTER))
    styles.add(ParagraphStyle('KPI_Label', parent=styles['Normal'],
                              fontSize=8, textColor=gray, alignment=TA_CENTER))
    styles.add(ParagraphStyle('Right_EA', parent=styles['Normal'],
                              fontSize=10, textColor=dark, alignment=TA_RIGHT))
    
    elements = []
    
    def fmt(val, suffix='EUR'):
        if suffix == 'EUR':
            return f"{val:,.0f} EUR".replace(',', '.')
        elif suffix == '%':
            return f"{val:.2f}%"
        return str(val)
    
    # HEADER
    elements.append(Paragraph("INVESTMENT EXPOSÉ", styles['Title_EA']))
    elements.append(Paragraph(
        f"10-Jahres-Prognose | Erstellt am {datetime.now(timezone.utc).strftime('%d.%m.%Y')}",
        styles['Subtitle_EA']
    ))
    
    # KPI ÜBERSICHT
    elements.append(Paragraph("Zusammenfassung", styles['Section_EA']))
    
    kpi_data = [
        ['Gesamtinvestition', 'Eigenkapital', 'Fremdkapital', 'IRR'],
        [fmt(result.total_investment), fmt(result.equity_invested),
         fmt(result.debt_amount), fmt(result.irr_percent, '%')],
        ['NPV', 'Eigenkapital-ROI', 'Gesamtgewinn', 'Break-Even'],
        [fmt(result.npv), fmt(result.equity_roi_percent, '%'),
         fmt(result.total_profit), f"Jahr {result.break_even_year}" if result.break_even_year else "–"]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[42*mm]*4)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), light_gray),
        ('BACKGROUND', (0, 2), (-1, 2), light_gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), gray),
        ('TEXTCOLOR', (0, 2), (-1, 2), gray),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 2), (-1, 2), 8),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('FONTSIZE', (0, 3), (-1, 3), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
        ('ROUNDEDCORNERS', [3, 3, 3, 3]),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 6*mm))
    
    # EINGABEPARAMETER
    elements.append(Paragraph("Eingabeparameter", styles['Section_EA']))
    
    param_data = [
        ['Parameter', 'Wert', 'Parameter', 'Wert'],
        ['Kaufpreis', fmt(inp.purchase_price), 'Mietsteigerung/Jahr', fmt(inp.rent_increase_percent, '%')],
        ['Renovierung', fmt(inp.renovation_costs), 'Wertsteigerung/Jahr', fmt(inp.appreciation_percent, '%')],
        ['Nebenkosten', fmt(inp.additional_costs_percent, '%'), 'Leerstandsrate', fmt(inp.vacancy_rate, '%')],
        ['Monatliche Miete', fmt(inp.monthly_rent), 'Betriebskosten', fmt(inp.running_costs_percent, '%')],
        ['Eigenkapital', fmt(inp.equity_percent, '%'), 'Hypothekenzins', fmt(inp.mortgage_rate, '%')],
        ['Haltedauer', f"{inp.holding_period} Jahre", 'Diskontierungszins', fmt(inp.discount_rate, '%')],
    ]
    
    param_table = Table(param_data, colWidths=[35*mm, 35*mm, 40*mm, 30*mm])
    param_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), dark),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), light_gray]),
    ]))
    elements.append(param_table)
    elements.append(Spacer(1, 6*mm))
    
    # 10-JAHRES-PROGNOSE TABELLE
    elements.append(Paragraph("10-Jahres-Cashflow-Prognose", styles['Section_EA']))
    
    cf_header = ['Jahr', 'Bruttomiete', 'Netto-Miete', 'Hypothek', 'Cashflow', 'Kum. CF', 'Immobilienwert']
    cf_rows = [cf_header]
    for y in result.yearly_data:
        cf_rows.append([
            str(y.year),
            fmt(y.gross_rent),
            fmt(y.net_rental_income),
            fmt(y.mortgage_payment),
            fmt(y.cashflow),
            fmt(y.cumulative_cashflow),
            fmt(y.property_value)
        ])
    
    col_w = [12*mm, 26*mm, 24*mm, 22*mm, 22*mm, 22*mm, 30*mm]
    cf_table = Table(cf_rows, colWidths=col_w)
    cf_style = [
        ('BACKGROUND', (0, 0), (-1, 0), dark),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#FFFFFF'), light_gray]),
    ]
    cf_table.setStyle(TableStyle(cf_style))
    elements.append(cf_table)
    elements.append(Spacer(1, 8*mm))
    
    # DISCLAIMER
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(
        "<b>WICHTIGER HINWEIS — KEINE ANLAGEBERATUNG</b>",
        ParagraphStyle('Disclaimer_Title', parent=styles['Normal'],
                      fontSize=9, textColor=HexColor('#B45309'), spaceBefore=4*mm, spaceAfter=2*mm)
    ))
    elements.append(Paragraph(
        "Diese Simulation dient ausschließlich zu Informations- und Veranschaulichungszwecken und stellt "
        "keine Anlageberatung, Kaufempfehlung oder Renditegarantie dar. Alle dargestellten Zahlen, Prognosen "
        "und Ergebnisse basieren auf den vom Benutzer eingegebenen Annahmen und vereinfachten Modellen. "
        "<b>Tatsächliche Ergebnisse können erheblich abweichen.</b>",
        styles['Small_EA']
    ))
    elements.append(Paragraph(
        "Insbesondere können Faktoren wie Steuern, Finanzierungskonditionen, Währungsrisiken, politische "
        "Veränderungen, Marktvolatilität, Instandhaltungskosten und unvorhergesehene Ereignisse die tatsächliche "
        "Rendite wesentlich beeinflussen. Konsultieren Sie vor jeder Investitionsentscheidung einen qualifizierten "
        "und unabhängigen Finanzberater. Der Herausgeber übernimmt keine Haftung für Entscheidungen, die auf "
        "Grundlage dieser Berechnung getroffen werden.",
        styles['Small_EA']
    ))
    
    # PDF generieren
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Investment_Expose.pdf"}
    )


# ---- DASHBOARD API ----

# =============================================
# CRM / PIPELINE API
# =============================================

def get_stage_probability(stage_id: str) -> int:
    for s in DEFAULT_PIPELINE_STAGES:
        if s["id"] == stage_id:
            return s["probability"]
    return 0

@api_router.get("/admin/crm/stages")
async def get_pipeline_stages(admin: str = Depends(verify_admin)):
    return DEFAULT_PIPELINE_STAGES

@api_router.get("/admin/crm/leads")
async def get_crm_leads(admin: str = Depends(verify_admin)):
    leads = await db.crm_leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return leads

@api_router.post("/admin/crm/leads")
async def create_crm_lead(lead: CRMLeadCreate, admin: str = Depends(verify_admin)):
    lead_dict = lead.model_dump()
    now = datetime.now(timezone.utc).isoformat()

    # Check if email already exists in CRM
    existing = await db.crm_leads.find_one({"email": lead_dict.get("email", "")})
    if existing:
        # Reactivate if lost
        if existing.get("status") == "lost":
            await db.crm_leads.update_one({"id": existing["id"]}, {"$set": {"status": "new"}})
        # Create new deal for existing lead
        await db.crm_deals.insert_one({
            "id": str(uuid.uuid4())[:8],
            "lead_id": existing["id"],
            "stage": "new_lead",
            "deal_value": 0,
            "probability": 10,
            "expected_revenue": 0,
            "assigned_to": None,
            "notes": None,
            "created_at": now,
            "updated_at": now
        })
        updated = await db.crm_leads.find_one({"id": existing["id"]}, {"_id": 0})
        return updated

    lead_dict["id"] = str(uuid.uuid4())[:8]
    lead_dict["status"] = "new"
    lead_dict["created_at"] = now
    await db.crm_leads.insert_one(lead_dict)

    # Auto-create deal in "New Lead" stage
    deal = {
        "id": str(uuid.uuid4())[:8],
        "lead_id": lead_dict["id"],
        "stage": "new_lead",
        "deal_value": 0,
        "probability": 10,
        "expected_revenue": 0,
        "assigned_to": None,
        "notes": None,
        "created_at": now,
        "updated_at": now
    }
    await db.crm_deals.insert_one(deal)

    created = await db.crm_leads.find_one({"id": lead_dict["id"]}, {"_id": 0})
    return created

@api_router.put("/admin/crm/leads/{lead_id}")
async def update_crm_lead(lead_id: str, lead: CRMLeadUpdate, admin: str = Depends(verify_admin)):
    update_data = {k: v for k, v in lead.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Keine Änderungen")
    result = await db.crm_leads.update_one({"id": lead_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead nicht gefunden")
    updated = await db.crm_leads.find_one({"id": lead_id}, {"_id": 0})
    return updated

@api_router.delete("/admin/crm/leads/{lead_id}")
async def delete_crm_lead(lead_id: str, admin: str = Depends(verify_admin)):
    await db.crm_deals.delete_many({"lead_id": lead_id})
    result = await db.crm_leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead nicht gefunden")
    return {"message": "Lead und zugehörige Deals gelöscht"}

# --- DEALS ---

@api_router.get("/admin/crm/deals")
async def get_crm_deals(admin: str = Depends(verify_admin)):
    deals = await db.crm_deals.find({}, {"_id": 0}).sort("updated_at", -1).to_list(1000)
    # Enrich with lead data
    for deal in deals:
        lead = await db.crm_leads.find_one({"id": deal.get("lead_id")}, {"_id": 0})
        deal["lead"] = lead
    return deals

@api_router.post("/admin/crm/deals")
async def create_crm_deal(deal: DealCreate, admin: str = Depends(verify_admin)):
    probability = get_stage_probability(deal.stage)
    deal_dict = deal.model_dump()
    deal_dict["id"] = str(uuid.uuid4())[:8]
    deal_dict["probability"] = probability
    deal_dict["expected_revenue"] = round(deal_dict["deal_value"] * probability / 100, 2)
    deal_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    deal_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.crm_deals.insert_one(deal_dict)
    created = await db.crm_deals.find_one({"id": deal_dict["id"]}, {"_id": 0})
    return created

@api_router.put("/admin/crm/deals/{deal_id}")
async def update_crm_deal(deal_id: str, deal: DealUpdate, admin: str = Depends(verify_admin)):
    update_data = {k: v for k, v in deal.model_dump().items() if v is not None}
    if "stage" in update_data:
        update_data["probability"] = get_stage_probability(update_data["stage"])
    # Get current deal to compute expected revenue
    current = await db.crm_deals.find_one({"id": deal_id}, {"_id": 0})
    if not current:
        raise HTTPException(status_code=404, detail="Deal nicht gefunden")
    new_value = update_data.get("deal_value", current.get("deal_value", 0))
    new_prob = update_data.get("probability", current.get("probability", 0))
    update_data["expected_revenue"] = round(new_value * new_prob / 100, 2)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.crm_deals.update_one({"id": deal_id}, {"$set": update_data})
    # Update lead status if deal won/lost
    if update_data.get("stage") == "won":
        await db.crm_leads.update_one({"id": current["lead_id"]}, {"$set": {"status": "won"}})
    elif update_data.get("stage") == "lost":
        await db.crm_leads.update_one({"id": current["lead_id"]}, {"$set": {"status": "lost"}})
    elif update_data.get("stage") and update_data["stage"] not in ["new_lead"]:
        await db.crm_leads.update_one({"id": current["lead_id"]}, {"$set": {"status": "qualified"}})

    updated = await db.crm_deals.find_one({"id": deal_id}, {"_id": 0})
    return updated

@api_router.delete("/admin/crm/deals/{deal_id}")
async def delete_crm_deal(deal_id: str, admin: str = Depends(verify_admin)):
    result = await db.crm_deals.delete_one({"id": deal_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Deal nicht gefunden")
    return {"message": "Deal gelöscht"}

# --- CRM DASHBOARD STATS ---

@api_router.get("/admin/crm/stats")
async def get_crm_stats(admin: str = Depends(verify_admin)):
    total_leads = await db.crm_leads.count_documents({})
    active_deals = await db.crm_deals.count_documents({"stage": {"$nin": ["won", "lost"]}})
    won_deals = await db.crm_deals.count_documents({"stage": "won"})
    lost_deals = await db.crm_deals.count_documents({"stage": "lost"})

    # Pipeline value & expected revenue
    pipeline = db.crm_deals.aggregate([
        {"$match": {"stage": {"$nin": ["lost"]}}},
        {"$group": {
            "_id": None,
            "pipeline_value": {"$sum": "$deal_value"},
            "expected_revenue": {"$sum": "$expected_revenue"},
            "won_revenue": {"$sum": {"$cond": [{"$eq": ["$stage", "won"]}, "$deal_value", 0]}}
        }}
    ])
    totals = await pipeline.to_list(1)
    totals = totals[0] if totals else {"pipeline_value": 0, "expected_revenue": 0, "won_revenue": 0}

    # By source
    source_pipeline = db.crm_leads.aggregate([
        {"$group": {"_id": "$lead_source", "count": {"$sum": 1}}}
    ])
    by_source = {doc["_id"]: doc["count"] async for doc in source_pipeline}

    # By stage
    stage_pipeline = db.crm_deals.aggregate([
        {"$group": {
            "_id": "$stage",
            "count": {"$sum": 1},
            "total_value": {"$sum": "$deal_value"},
            "total_expected": {"$sum": "$expected_revenue"}
        }}
    ])
    by_stage = {}
    async for doc in stage_pipeline:
        by_stage[doc["_id"]] = {"count": doc["count"], "value": doc["total_value"], "expected": doc["total_expected"]}

    # Conversion rate
    conversion_rate = round((won_deals / total_leads * 100), 1) if total_leads > 0 else 0

    return {
        "total_leads": total_leads,
        "active_deals": active_deals,
        "won_deals": won_deals,
        "lost_deals": lost_deals,
        "pipeline_value": totals.get("pipeline_value", 0),
        "expected_revenue": totals.get("expected_revenue", 0),
        "won_revenue": totals.get("won_revenue", 0),
        "conversion_rate": conversion_rate,
        "by_source": by_source,
        "by_stage": by_stage
    }

# --- MIGRATE EXISTING LEADS ---

@api_router.post("/admin/crm/migrate")
async def migrate_existing_leads(admin: str = Depends(verify_admin)):
    """Migrate existing leads from old leads collection to CRM"""
    old_leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    contacts = await db.contact_submissions.find({}, {"_id": 0}).to_list(1000)
    migrated = 0

    # Existing CRM leads by email
    existing = await db.crm_leads.find({}, {"_id": 0, "email": 1, "id": 1}).to_list(1000)
    existing_map = {l["email"]: l["id"] for l in existing if l.get("email")}
    seen_emails = set()

    for lead in old_leads:
        email = lead.get("email", "")
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)
        now = lead.get("submitted_at", datetime.now(timezone.utc).isoformat())

        if email in existing_map:
            # Lead exists — add new deal
            await db.crm_deals.insert_one({
                "id": str(uuid.uuid4())[:8],
                "lead_id": existing_map[email],
                "stage": "new_lead",
                "deal_value": 0, "probability": 10, "expected_revenue": 0,
                "assigned_to": None, "notes": None,
                "created_at": now, "updated_at": now
            })
            await db.crm_leads.update_one({"id": existing_map[email]}, {"$set": {"status": "new"}})
        else:
            crm_lead = {
                "id": str(uuid.uuid4())[:8],
                "name": lead.get("name", ""),
                "email": email,
                "phone": lead.get("phone"),
                "lead_source": lead.get("source", "website"),
                "utm_source": None, "utm_medium": None, "utm_campaign": None,
                "entry_page": None,
                "tool_used": lead.get("type", lead.get("source", "unknown")),
                "status": "new",
                "created_at": now
            }
            await db.crm_leads.insert_one(crm_lead)
            await db.crm_deals.insert_one({
                "id": str(uuid.uuid4())[:8],
                "lead_id": crm_lead["id"],
                "stage": "new_lead",
                "deal_value": 0, "probability": 10, "expected_revenue": 0,
                "assigned_to": None, "notes": None,
                "created_at": now, "updated_at": now
            })
            existing_map[email] = crm_lead["id"]
        migrated += 1

    for contact in contacts:
        email = contact.get("email", "")
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)
        now = contact.get("submitted_at", datetime.now(timezone.utc).isoformat())

        if email in existing_map:
            await db.crm_deals.insert_one({
                "id": str(uuid.uuid4())[:8],
                "lead_id": existing_map[email],
                "stage": "new_lead",
                "deal_value": 0, "probability": 10, "expected_revenue": 0,
                "assigned_to": None, "notes": f"Betreff: {contact.get('subject', '')}",
                "created_at": now, "updated_at": now
            })
            await db.crm_leads.update_one({"id": existing_map[email]}, {"$set": {"status": "new"}})
        else:
            crm_lead = {
                "id": str(uuid.uuid4())[:8],
                "name": contact.get("name", ""),
                "email": email,
                "phone": contact.get("phone"),
                "lead_source": "kontaktformular",
                "utm_source": None, "utm_medium": None, "utm_campaign": None,
                "entry_page": None,
                "tool_used": "contact_form",
                "status": "new",
                "created_at": now
            }
            await db.crm_leads.insert_one(crm_lead)
            await db.crm_deals.insert_one({
                "id": str(uuid.uuid4())[:8],
                "lead_id": crm_lead["id"],
                "stage": "new_lead",
                "deal_value": 0, "probability": 10, "expected_revenue": 0,
                "assigned_to": None, "notes": f"Betreff: {contact.get('subject', '')}",
                "created_at": now, "updated_at": now
            })
            existing_map[email] = crm_lead["id"]
        migrated += 1

    return {"migrated": migrated, "total_crm_leads": await db.crm_leads.count_documents({})}

@api_router.delete("/admin/crm/reset")
async def reset_crm_data(admin: str = Depends(verify_admin)):
    """Reset all CRM data (leads + deals)"""
    deleted_deals = await db.crm_deals.delete_many({})
    deleted_leads = await db.crm_leads.delete_many({})
    return {
        "message": "CRM-Daten zurückgesetzt",
        "deleted_leads": deleted_leads.deleted_count,
        "deleted_deals": deleted_deals.deleted_count
    }


# --- AUTO-LINK NEW CONTACT FORM LEADS TO CRM ---

# =============================================
# EVENTS API
# =============================================

# =============================================
# LEISTUNGEN CMS API
# =============================================

LEISTUNGEN_DEFAULTS = {
    "hero": {
        "tagline": "Präzise. Effizient. Strategisch.",
        "title": "Unsere Leistungen für Ihre Zukunft",
        "description": "Wir entwickeln Konzepte für die Unternehmensgründung, Vermögensstrukturierung, Immobilien, Aufenthalts- und Finanzlösungen — genau auf Ihre Situation zugeschnitten."
    },
    "services": [
        {
            "id": "immobilien",
            "title": "Immobilien & Aufenthalt",
            "tagline": "Finden. Sichern. Leben.",
            "description": "Finden und sichern Sie Ihre Traumimmobilie und regeln Sie Ihren Aufenthaltsstatus unkompliziert.",
            "image": "",
            "points": ["Immobiliensuche & Due Diligence", "Aufenthaltsgenehmigung & Visa", "Katasterprüfung & Eigentumsübertragung", "Mietrendite-Optimierung"]
        },
        {
            "id": "gruendung",
            "title": "Unternehmensgründung & Strukturierung",
            "tagline": "Gründen. Strukturieren. Wachsen.",
            "description": "Wir gestalten Ihre internationale Unternehmensarchitektur für maximale Effizienz und Rechtssicherheit.",
            "image": "",
            "points": ["Firmengründung in Montenegro (9% KSt)", "Holding-Strukturen & Steueroptimierung", "Bankkonto-Eröffnung & KYC-Begleitung", "Geschäftsadresse & Virtual Office"]
        },
        {
            "id": "legal",
            "title": "Legal- & Finanzdienstleistungen",
            "tagline": "Prüfen. Absichern. Durchsetzen.",
            "description": "Führen Sie Ihr Vorhaben sicher durch die rechtlichen und finanziellen Anforderungen vor Ort.",
            "image": "",
            "points": ["Vertragsgestaltung & -prüfung", "Steuerberatung Montenegro & Serbien", "AML/KYC Compliance", "Streitbeilegung & Mediation"]
        },
        {
            "id": "investor",
            "title": "Investor Relations & Projektvermittlung",
            "tagline": "Verbinden. Vermitteln. Realisieren.",
            "description": "Wir verbinden Sie mit exklusiven Off-Market Investitionsmöglichkeiten und lokalen Partnern.",
            "image": "",
            "points": ["Off-Market Deal-Zugang", "Investoren-Matching & Co-Investments", "Projektentwicklung & Feasibility", "Exit-Strategien & Wiederverkauf"]
        }
    ],
    "legal_risks": {
        "title": "Juristische Stabilisierung",
        "description": "Das größte Risiko in Montenegro sind Immobilien, die zwar physisch existieren, aber formal nicht legalisiert, nicht abgenommen oder nicht nutzungsfähig sind.",
        "conclusion": "Ihr Vorteil: Aus einem juristischen Risiko wird ein rechtlich abgesichertes, werthaltiges Asset.",
        "items": [
            {"id": "baugenehmigung", "problem": "Fehlende Baugenehmigung & Legalisierung", "risk": "Drohendes Nutzungsverbot, Abrissverfügung oder massiver Wertabschlag bei nicht legalisierten Objekten.", "solution": "EuroAdria Corporate Solutions: Vollständige Legalisierungsprüfung und Nachholung aller behördlichen Genehmigungen durch unser Netzwerk."},
            {"id": "kataster", "problem": "Unklare Katastereinträge", "risk": "Der Verkäufer steht noch im Kataster, alte Grundschulden oder Dienstbarkeiten sind nicht gelöscht.", "solution": "Forensische Katasterrecherche bis 1945. Bereinigung aller Altlasten vor Vertragsunterzeichnung."},
            {"id": "erbchaos", "problem": "Erbchaos & Familienkonflikte", "risk": "Miteigentümer oder Dritte stellen den Deal nachträglich infrage und schaffen Erpressungspotenzial.", "solution": "Lückenlose Eigentümerkette. Notarielle Absicherung aller Parteien. Keine offenen Ansprüche."}
        ]
    },
    "compliance_risks": {
        "title": "Exit-Sicherheit & Compliance",
        "description": "Renditeobjekte erfordern mehr als nur ein schönes Gebäude. Sie erfordern klare Compliance und eine steuerlich verteidigbare Zahlungsstruktur.",
        "conclusion": "Kein politischer Blindflug — garantierte Wiederverkaufsfähigkeit Ihres Objekts.",
        "items": [
            {"id": "airbnb", "problem": "Illegale Kurzzeitvermietung (Airbnb)", "risk": "Viele Objekte sind nicht für touristische Nutzung zugelassen — das gefährdet Ihre gesamte Rendite.", "solution": "Prüfung & Beantragung der touristischen Nutzungsgenehmigung. Registrierung bei der Steuerbehörde."},
            {"id": "grauzonen", "problem": "Grauzonen-Zahlungen", "risk": "Niedriger Kaufpreis in der Urkunde, Rest bar — zerstört die Exit-Strategie und schafft ein Geldwäschethema.", "solution": "Transparente Kaufpreisstruktur. Vollständige Dokumentation für westliche Banken und Steuerbehörden."},
            {"id": "politisch", "problem": "Politische & regulatorische Risiken", "risk": "Plötzlicher Baustopp, Enteignung oder Nutzungsänderung durch die Gemeinde.", "solution": "Politisches Screening vor Ort. Frühwarnsystem durch unser Behörden-Netzwerk in Montenegro."}
        ]
    },
    "guarantee": {
        "title": "Die doppelte Garantie",
        "subtitle": "Sicherheit für Neukäufer, Sanierung für Bestandsbesitzer",
        "buyer": {
            "label": "VOR DEM KAUF",
            "title": "Für den Erstkäufer",
            "description": "Wir erstellen ein vollständiges Compliance-Dossier zur Immobilie, das alle Risikobereiche abdeckt: Legalisierung, Kataster, Standort-Screening und Exitfähigkeit.",
            "highlight": "Wir garantieren Klarheit und Compliance, bevor Sie den Kaufvertrag unterzeichnen. Keine versteckten Risiken, keine unkalkulierbaren Überraschungen."
        },
        "owner": {
            "label": "NACH DEM KAUF",
            "title": "Für den Bestandsbesitzer",
            "description": "Wenn Ihr Investment bereits von juristischen Mängeln betroffen ist, übernehmen wir die Schadensbegrenzung und juristische Sanierung Ihres Objekts.",
            "highlight": "Wir bereinigen offene Ansprüche, führen fehlende Eintragungen durch und machen aus einem Risiko-Objekt ein bankfähiges, exitfähiges Asset."
        }
    },
    "cta": {
        "title": "Kontaktieren Sie uns für eine Erstberatung",
        "description": "Sie planen den Kauf oder sitzen bereits in der Risikozone? Nehmen Sie jetzt Kontakt mit unserer Kanzlei-gestützten Task Force auf.",
        "button_text": "Jetzt unverbindliche Erstberatung anfragen"
    }
}

@api_router.get("/leistungen-content")
async def get_leistungen_content():
    content = await db.leistungen_content.find_one({"key": "main"}, {"_id": 0})
    if not content:
        return LEISTUNGEN_DEFAULTS
    return {k: v for k, v in content.items() if k != "key"}

@api_router.put("/admin/leistungen-content")
async def update_leistungen_content(data: dict, admin: str = Depends(verify_admin)):
    data["key"] = "main"
    await db.leistungen_content.update_one(
        {"key": "main"}, {"$set": data}, upsert=True
    )
    updated = await db.leistungen_content.find_one({"key": "main"}, {"_id": 0})
    return {k: v for k, v in updated.items() if k != "key"}

@api_router.get("/events")
async def get_events():
    events = await db.events.find({}, {"_id": 0}).sort("date", 1).to_list(100)
    return events

@api_router.get("/events/{event_id}")
async def get_event(event_id: str):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event nicht gefunden")
    return event

@api_router.post("/admin/events")
async def create_event(event: EventCreate, admin: str = Depends(verify_admin)):
    event_dict = event.model_dump()
    event_dict["id"] = str(uuid.uuid4())[:8]
    event_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.events.insert_one(event_dict)
    created = await db.events.find_one({"id": event_dict["id"]}, {"_id": 0})
    return created

@api_router.put("/admin/events/{event_id}")
async def update_event(event_id: str, event: EventUpdate, admin: str = Depends(verify_admin)):
    update_data = {k: v for k, v in event.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Keine Änderungen")
    result = await db.events.update_one({"id": event_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event nicht gefunden")
    updated = await db.events.find_one({"id": event_id}, {"_id": 0})
    return updated

@api_router.delete("/admin/events/{event_id}")
async def delete_event(event_id: str, admin: str = Depends(verify_admin)):
    result = await db.events.delete_one({"id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event nicht gefunden")
    return {"message": "Event gelöscht"}


@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    """Get investment dashboard statistics"""
    locations = await db.locations.find({}, {"_id": 0}).to_list(1000)
    projects = await db.infrastructure_projects.find({}, {"_id": 0}).to_list(1000)
    zones = await db.opportunity_zones.find({}, {"_id": 0}).to_list(1000)
    
    # Recalculate scores
    for loc in locations:
        infra_boost = await calculate_infrastructure_boost(loc["latitude"], loc["longitude"])
        adjusted_infra = min(loc["infrastructure_score"] + infra_boost, 100)
        loc["investment_score"] = calculate_investment_score(
            adjusted_infra,
            loc["tourism_growth"],
            loc["rental_yield"],
            loc["price_growth"]
        )
    
    # Sort for rankings
    by_score = sorted(locations, key=lambda x: x["investment_score"], reverse=True)
    by_yield = sorted(locations, key=lambda x: x["rental_yield"], reverse=True)
    by_growth = sorted(locations, key=lambda x: x["price_growth"], reverse=True)
    
    # Filter new infrastructure projects
    new_projects = [p for p in projects if p["status"] in ["construction", "planned"]]
    
    return {
        "total_locations": len(locations),
        "total_infrastructure": len(projects),
        "total_zones": len(zones),
        "top_investment": by_score[:5],
        "highest_yield": by_yield[:5],
        "strongest_growth": by_growth[:5],
        "new_infrastructure": new_projects[:5],
        "countries": {
            "Montenegro": len([l for l in locations if l["country"] == "Montenegro"]),
            "Serbien": len([l for l in locations if l["country"] == "Serbien"])
        }
    }


# ---- SEED DATA API ----

@api_router.post("/admin/seed-investment-data")
async def seed_investment_data(admin: str = Depends(verify_admin)):
    """Seed the database with initial investment data"""
    
    # Seed Locations
    loc_count = await db.locations.count_documents({})
    if loc_count == 0:
        for loc in SEED_LOCATIONS:
            loc_dict = loc.copy()
            loc_dict["id"] = str(uuid.uuid4())
            loc_dict["investment_score"] = calculate_investment_score(
                loc["infrastructure_score"],
                loc["tourism_growth"],
                loc["rental_yield"],
                loc["price_growth"]
            )
            loc_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.locations.insert_one(loc_dict)
    
    # Seed Infrastructure
    infra_count = await db.infrastructure_projects.count_documents({})
    if infra_count == 0:
        for proj in SEED_INFRASTRUCTURE:
            proj_dict = proj.copy()
            proj_dict["id"] = str(uuid.uuid4())
            proj_dict["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.infrastructure_projects.insert_one(proj_dict)
    
    # Seed Zones
    zone_count = await db.opportunity_zones.count_documents({})
    if zone_count == 0:
        for zone in SEED_ZONES:
            zone_dict = zone.copy()
            zone_dict["id"] = str(uuid.uuid4())
            await db.opportunity_zones.insert_one(zone_dict)
    
    return {
        "message": "Investment data seeded successfully",
        "locations": len(SEED_LOCATIONS),
        "infrastructure": len(SEED_INFRASTRUCTURE),
        "zones": len(SEED_ZONES)
    }


# ─── File Upload (Object Storage) ───────────────────────────────────────
@api_router.post("/admin/storage/upload")
async def admin_upload_file(file: UploadFile = File(...), admin: str = Depends(verify_admin)):
    """Upload image or file to object storage, returns public serve URL"""
    allowed_ext = {"jpg", "jpeg", "png", "gif", "webp", "pdf", "csv", "txt"}
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"Dateityp .{ext} nicht erlaubt")
    
    mime_map = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
        "gif": "image/gif", "webp": "image/webp", "pdf": "application/pdf",
        "csv": "text/csv", "txt": "text/plain"
    }
    content_type = file.content_type or mime_map.get(ext, "application/octet-stream")
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/uploads/{file_id}.{ext}"
    
    data = await file.read()
    result = put_object(storage_path, data, content_type)
    
    await db.uploaded_files.insert_one({
        "file_id": file_id,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": content_type,
        "size": result.get("size", len(data)),
        "ext": ext,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "file_id": file_id,
        "filename": file.filename,
        "url": f"/api/files/{result['path']}",
        "content_type": content_type
    }

@api_router.get("/files/{path:path}")
async def serve_file(path: str):
    """Serve uploaded file from object storage"""
    record = await db.uploaded_files.find_one({"storage_path": path}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Datei nicht gefunden")
    data, ct = get_object(path)
    return Response(content=data, media_type=record.get("content_type", ct))



# ---- YOUTUBE API ----
_youtube_cache = {"data": None, "expires": 0}

@api_router.get("/youtube/latest")
async def get_youtube_latest():
    now = time.time()
    if _youtube_cache["data"] and now < _youtube_cache["expires"]:
        return _youtube_cache["data"]

    api_key = os.environ.get("YOUTUBE_API_KEY", "")
    playlist_id = os.environ.get("YOUTUBE_PLAYLIST_ID", "UULJ7QhMlsoV23wgYusibA3A")

    if not api_key:
        raise HTTPException(status_code=500, detail="YouTube API key not configured")

    try:
        url = "https://www.googleapis.com/youtube/v3/playlistItems"
        params = {
            "part": "snippet",
            "playlistId": playlist_id,
            "maxResults": 12,
            "key": api_key
        }
        resp = http_requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        raw = resp.json()

        videos = []
        for item in raw.get("items", []):
            s = item.get("snippet", {})
            vid_id = s.get("resourceId", {}).get("videoId")
            if not vid_id:
                continue
            videos.append({
                "id": vid_id,
                "title": s.get("title", ""),
                "thumbnail": s.get("thumbnails", {}).get("high", {}).get("url", f"https://img.youtube.com/vi/{vid_id}/hqdefault.jpg"),
                "publishedAt": s.get("publishedAt", ""),
            })

        # Fetch view counts via videos endpoint
        if videos:
            vid_ids = ",".join(v["id"] for v in videos)
            stats_url = "https://www.googleapis.com/youtube/v3/videos"
            stats_params = {"part": "statistics", "id": vid_ids, "key": api_key}
            stats_resp = http_requests.get(stats_url, params=stats_params, timeout=10)
            if stats_resp.ok:
                stats_data = stats_resp.json()
                view_map = {}
                for si in stats_data.get("items", []):
                    vc = int(si.get("statistics", {}).get("viewCount", 0))
                    if vc >= 1000:
                        view_map[si["id"]] = f"{vc / 1000:.1f}K"
                    else:
                        view_map[si["id"]] = str(vc)
                for v in videos:
                    v["views"] = view_map.get(v["id"], "0")

        result = {"videos": videos, "channelUrl": "https://youtube.com/@euroadriacs"}
        _youtube_cache["data"] = result
        _youtube_cache["expires"] = now + 3600  # Cache 1 hour
        return result

    except http_requests.exceptions.RequestException as e:
        if _youtube_cache["data"]:
            return _youtube_cache["data"]
        raise HTTPException(status_code=502, detail=f"YouTube API error: {str(e)}")


# ---- TRANSLATION API ----
class TranslateRequest(BaseModel):
    text: str
    source: str = "de"
    target: str = "en"

class TranslateBatchRequest(BaseModel):
    texts: List[str]
    source: str = "de"
    target: str = "en"

# Try to import argostranslate (available in dev, not on Render)
_argos_available = False
_argos_de_en = None
_argos_en_de = None

try:
    from argostranslate import package as argos_package, translate as argos_translate
    installed = argos_translate.get_installed_languages()
    lang_codes = {l.code: l for l in installed}
    if "de" in lang_codes and "en" in lang_codes:
        _argos_de_en = lang_codes["de"].get_translation(lang_codes["en"])
        _argos_en_de = lang_codes["en"].get_translation(lang_codes["de"])
        if _argos_de_en and _argos_en_de:
            _argos_available = True
except Exception:
    pass


def _translate_text_argos(text: str, source: str, target: str) -> str:
    if source == "de" and target == "en" and _argos_de_en:
        return _argos_de_en.translate(text)
    elif source == "en" and target == "de" and _argos_en_de:
        return _argos_en_de.translate(text)
    return text


def _translate_text_api(text: str, source: str, target: str) -> str:
    """Fallback: use MyMemory free API (EU-based, DSGVO-friendly)"""
    import re
    
    # For short texts, translate directly
    if len(text) <= 450:
        try:
            resp = http_requests.get(
                "https://api.mymemory.translated.net/get",
                params={"q": text, "langpair": f"{source}|{target}"},
                timeout=10
            )
            if resp.ok:
                data = resp.json()
                translated = data.get("responseData", {}).get("translatedText", "")
                if translated and translated.upper() != text.upper():
                    return translated
        except Exception:
            pass
        return text

    # For HTML content: split by paragraphs/headings, translate each piece
    # Split on HTML block elements while keeping the tags
    html_pattern = r'(<(?:p|h[1-6]|li|div|blockquote|td|th|caption|figcaption)[^>]*>)(.*?)(</(?:p|h[1-6]|li|div|blockquote|td|th|caption|figcaption)>)'
    
    def translate_chunk(chunk):
        """Translate a single chunk, handling the 450 char limit"""
        if not chunk or not chunk.strip():
            return chunk
        # Strip HTML tags for translation
        clean = re.sub(r'<[^>]+>', '', chunk).strip()
        if not clean:
            return chunk
        if len(clean) <= 450:
            try:
                resp = http_requests.get(
                    "https://api.mymemory.translated.net/get",
                    params={"q": clean, "langpair": f"{source}|{target}"},
                    timeout=10
                )
                if resp.ok:
                    data = resp.json()
                    tr = data.get("responseData", {}).get("translatedText", "")
                    if tr and tr.upper() != clean.upper():
                        # Replace the text content in the original chunk (keep inline HTML)
                        return chunk.replace(clean, tr)
            except Exception:
                pass
            return chunk
        # Split longer text by sentences
        sentences = re.split(r'(?<=[.!?])\s+', clean)
        translated_sentences = []
        for sent in sentences:
            if len(sent) > 450:
                translated_sentences.append(sent)
                continue
            try:
                resp = http_requests.get(
                    "https://api.mymemory.translated.net/get",
                    params={"q": sent, "langpair": f"{source}|{target}"},
                    timeout=10
                )
                if resp.ok:
                    data = resp.json()
                    tr = data.get("responseData", {}).get("translatedText", "")
                    translated_sentences.append(tr if tr and tr.upper() != sent.upper() else sent)
                else:
                    translated_sentences.append(sent)
            except Exception:
                translated_sentences.append(sent)
        translated_text = " ".join(translated_sentences)
        return chunk.replace(clean, translated_text)
    
    # Try HTML-aware splitting first
    matches = list(re.finditer(html_pattern, text, re.DOTALL | re.IGNORECASE))
    if matches:
        result = text
        for match in reversed(matches):  # Reverse to preserve offsets
            full = match.group(0)
            inner = match.group(2)
            translated_inner = translate_chunk(inner)
            if translated_inner != inner:
                result = result[:match.start()] + match.group(1) + translated_inner + match.group(3) + result[match.end():]
        return result
    
    # Fallback: split by newlines/paragraphs
    parts = re.split(r'\n\n+', text)
    translated_parts = []
    for part in parts:
        translated_parts.append(translate_chunk(part))
    return "\n\n".join(translated_parts)


async def _get_or_translate(text: str, source: str, target: str) -> str:
    if not text or not text.strip():
        return text
    if source == target:
        return text

    text_hash = hashlib.md5(f"{text}:{source}:{target}".encode()).hexdigest()

    # Check MongoDB cache
    cached = await db.translations.find_one(
        {"hash": text_hash}, {"_id": 0, "translated": 1}
    )
    if cached:
        return cached["translated"]

    # Translate
    if _argos_available:
        translated = _translate_text_argos(text, source, target)
    else:
        translated = _translate_text_api(text, source, target)

    # Cache in MongoDB
    if translated and translated != text:
        await db.translations.update_one(
            {"hash": text_hash},
            {"$set": {
                "hash": text_hash,
                "original": text,
                "translated": translated,
                "source": source,
                "target": target,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )

    return translated


@api_router.post("/translate")
async def translate_text(req: TranslateRequest):
    translated = await _get_or_translate(req.text, req.source, req.target)
    return {"translatedText": translated, "source": req.source, "target": req.target}


@api_router.post("/translate/batch")
async def translate_batch(req: TranslateBatchRequest):
    results = []
    for text in req.texts[:50]:  # Limit to 50 texts per batch
        translated = await _get_or_translate(text, req.source, req.target)
        results.append(translated)
    return {"translations": results, "source": req.source, "target": req.target}


@api_router.get("/translate/article/{slug}")
async def translate_article(slug: str, target: str = "en"):
    article = await db.articles.find_one({"slug": slug}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")

    source = "de"
    if target == source:
        return article

    # Translate key fields
    for field in ["title", "excerpt", "content"]:
        if article.get(field):
            article[field] = await _get_or_translate(article[field], source, target)

    # Translate expert tip
    if article.get("expertTip"):
        for f in ["title", "content"]:
            if article["expertTip"].get(f):
                article["expertTip"][f] = await _get_or_translate(article["expertTip"][f], source, target)

    # Translate due diligence box
    if article.get("dueDiligenceBox"):
        if article["dueDiligenceBox"].get("title"):
            article["dueDiligenceBox"]["title"] = await _get_or_translate(article["dueDiligenceBox"]["title"], source, target)
        if article["dueDiligenceBox"].get("items"):
            items = []
            for item in article["dueDiligenceBox"]["items"]:
                items.append(await _get_or_translate(item, source, target))
            article["dueDiligenceBox"]["items"] = items

    return article


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'https://invest.euroadria.me,https://euroadria.me').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@app.on_event("startup")
async def startup_log():
    logger.info("Server started - Object storage will init on first use")